from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy.exc import ProgrammingError
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, date, timedelta
from functools import wraps
import os
import hashlib
import random
import string
import socket
import unicodedata
import secrets
import shutil

app = Flask(__name__)

# Segurança: configurações seguras por padrão
app.config.setdefault('DEBUG', False)
app.config.setdefault('SESSION_COOKIE_HTTPONLY', True)
# Em produção (quando a variável ENV=production estiver definida), forçar cookies seguros
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('ENV', '').lower() == 'production'
app.config.setdefault('PREFERRED_URL_SCHEME', 'https')


@app.after_request
def set_secure_headers(response):
    # Segurança básica de headers HTTP
    response.headers.setdefault('X-Content-Type-Options', 'nosniff')
    response.headers.setdefault('X-Frame-Options', 'DENY')
    response.headers.setdefault('Referrer-Policy', 'no-referrer-when-downgrade')
    response.headers.setdefault('Permissions-Policy', "accelerometer=(), camera=(), microphone=(), geolocation=()")
    # HSTS apenas em produção
    if os.environ.get('ENV', '').lower() == 'production':
        response.headers.setdefault('Strict-Transport-Security', 'max-age=63072000; includeSubDomains; preload')
    # CSP mínima (permitir recursos do mesmo host e CDNs usados)
    csp = (
        "default-src 'self' 'unsafe-inline' https: data:; "
        "script-src 'self' 'unsafe-inline' "
        "https://cdn.jsdelivr.net https://code.jquery.com https://cdn.datatables.net https://unpkg.com https://cdnjs.cloudflare.com; "
        "style-src 'self' 'unsafe-inline' "
        "https://cdn.jsdelivr.net https://cdn.datatables.net https://unpkg.com https://cdnjs.cloudflare.com; "
        "img-src 'self' https: data:; "
        "font-src 'self' https://cdnjs.cloudflare.com https://cdn.jsdelivr.net;"
    )
    response.headers.setdefault('Content-Security-Policy', csp)
    return response

def _load_or_create_secret_key():
    # Prefer env var for production deployments
    env_key = os.environ.get('SECRET_KEY')
    if env_key and env_key.strip():
        return env_key.strip()

    # Persist a stable key locally to avoid invalidating sessions on restart
    basedir_local = os.path.abspath(os.path.dirname(__file__))
    instance_dir = os.path.join(basedir_local, 'instance')
    os.makedirs(instance_dir, exist_ok=True)
    key_file = os.path.join(instance_dir, 'secret_key.txt')

    try:
        if os.path.exists(key_file):
            with open(key_file, 'r', encoding='utf-8') as f:
                existing = (f.read() or '').strip()
                if existing:
                    return existing
    except Exception:
        pass

    new_key = secrets.token_urlsafe(48)
    try:
        with open(key_file, 'w', encoding='utf-8') as f:
            f.write(new_key)
    except Exception:
        # Fallback: still run, but sessions will reset on restart
        pass
    return new_key

app.config['SECRET_KEY'] = _load_or_create_secret_key()

# ==================== MIDDLEWARE DE VERIFICAÃ‡ÃƒO DE LICENÃ‡A ====================

# Define o caminho absoluto para o banco de dados
basedir = os.path.abspath(os.path.dirname(__file__))
db_path = os.path.join(basedir, 'database')
# Criar pasta database se nÃ£o existir
os.makedirs(db_path, exist_ok=True)
# Banco SQLite versionado/esperado em `database/caixa.db` (dev/local).
_sqlite_db_file = os.path.join(db_path, 'caixa.db')
# Render/Prod: usar DATABASE_URL (PostgreSQL). Dev local: SQLite.
# Railway/Render podem usar nomes diferentes de env.
database_url = (
    os.environ.get('DATABASE_URL')
    or os.environ.get('POSTGRES_URL')
    or os.environ.get('POSTGRES_PRISMA_URL')
    or os.environ.get('POSTGRES_URL_NON_POOLING')
)
if database_url:
    database_url = database_url.strip().strip('"').strip("'")
if database_url:
    if database_url.startswith('postgres://'):
        database_url = database_url.replace('postgres://', 'postgresql://', 1)
    if database_url.startswith('postgresql+psycopg2://'):
        database_url = database_url.replace('postgresql+psycopg2://', 'postgresql+psycopg://', 1)
    elif database_url.startswith('postgresql://'):
        database_url = database_url.replace('postgresql://', 'postgresql+psycopg://', 1)
    # Se DATABASE_URL estiver inválida (ex.: role sem LOGIN), cair para SQLite ao invés de quebrar o app.
    try:
        from sqlalchemy import create_engine
        test_url = database_url
        if 'connect_timeout=' not in test_url:
            test_url = test_url + ('&' if '?' in test_url else '?') + 'connect_timeout=5'
        with create_engine(test_url).connect():
            pass
        app.config['SQLALCHEMY_DATABASE_URI'] = database_url
    except Exception as e:
        print(f"[WARN] Falha ao conectar no PostgreSQL via DATABASE_URL: {e}. Usando SQLite local.")
        app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(db_path, 'caixa.db')
else:
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(db_path, 'caixa.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# ==================== DB INIT (AUTO) ====================

def _ensure_db_ready():
    with app.app_context():
        db.create_all()
        try:
            from sqlalchemy import text
            with db.engine.connect() as conn:
                conn.execute(text('ALTER TABLE compra ADD COLUMN observacao TEXT'))
                conn.commit()
        except Exception:
            pass
        # CorreÃ§Ã£o automÃ¡tica: Adicionar coluna acesso_relatorios se nÃ£o existir
        try:
            from sqlalchemy import text
            with db.engine.connect() as conn:
                conn.execute(text('ALTER TABLE usuario ADD COLUMN acesso_relatorios BOOLEAN DEFAULT 0'))
                conn.commit()
        except Exception:
            pass
        try:
            from sqlalchemy import text
            with db.engine.connect() as conn:
                conn.execute(text('ALTER TABLE produto ADD COLUMN tamanho VARCHAR(100)'))
                conn.commit()
        except Exception:
            pass
        # Criar registros padrÃ£o (inclui ADMIN MASTER e ADMIN)
        try:
            init_db()
        except Exception:
            pass

# Executar no startup para evitar erro de tabela inexistente no primeiro request
try:
    _ensure_db_ready()
except Exception:
    pass

_db_initialized = False

@app.before_request
def ensure_db_initialized():
    global _db_initialized
    if _db_initialized:
        return
    try:
        _ensure_db_ready()
        _db_initialized = True
    except Exception:
        pass




# ==================== CONTEXT PROCESSORS ====================

@app.context_processor
def inject_user():
    """Injeta informaÃ§Ãµes do usuÃ¡rio em todos os templates"""
    if 'user_id' in session:
        usuario = db.session.get(Usuario, session['user_id'])
        return dict(usuario_logado=usuario)
    return dict(usuario_logado=None)

@app.context_processor
def inject_datetime():
    from datetime import datetime
    return dict(datetime=datetime)

# No contexto do processador, adicione a funÃ§Ã£o now()
@app.context_processor
def utility_processor():
    """Injeta funÃ§Ãµes utilitÃ¡rias em todos os templates"""
    def calcular_totais_caixa_template(caixa):
        return calcular_totais_caixa(caixa)
    def now():
        return datetime.utcnow()
    def format_currency(value):
        return formatar_moeda(value)
    return dict(
        calcular_totais_caixa=calcular_totais_caixa_template, 
        now=now,
        format_currency=format_currency
    )


# Adicione esta funÃ§Ã£o no seu app.py
def formatar_moeda(valor):
    """Formata valor para moeda brasileira R$ 1.234,56"""
    if valor is None:
        return "R$ 0,00"
    try:
        return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return f"R$ {valor}"

def parse_moeda(valor, default=0.0):
    """Converte string moeda BR (ex: 1.234,56) para float"""
    if valor is None:
        return default
    if isinstance(valor, (int, float)):
        return float(valor)
    s = str(valor).strip()
    if s == '':
        return default
    s = s.replace('R$', '').replace(' ', '')
    # Detect format:
    # - If contains comma, assume Brazilian format: '.' thousands, ',' decimal -> remove thousands and convert decimal to dot
    # - If contains no comma but contains dot, assume dot is decimal separator (en format) -> keep dot
    # - Otherwise plain integer string
    if ',' in s:
        s = s.replace('.', '').replace(',', '.')
    else:
        # keep dots as decimal separator
        s = s
    try:
        return float(s)
    except Exception:
        return default



# ==================== MODELS ====================

class Usuario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    senha = db.Column(db.String(200), nullable=False)
    perfil = db.Column(db.String(20), default='OPERADOR')
    acesso_dashboard = db.Column(db.Boolean, default=True)
    acesso_configuracoes = db.Column(db.Boolean, default=False)
    acesso_relatorios = db.Column(db.Boolean, default=False)  # NOVO CAMPO
    ativo = db.Column(db.Boolean, default=True)
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)

class Caixa(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    data = db.Column(db.Date, nullable=False)
    turno = db.Column(db.String(20), nullable=False)
    operador_id = db.Column(db.Integer, db.ForeignKey('usuario.id'))
    saldo_inicial = db.Column(db.Float, default=0)
    saldo_final = db.Column(db.Float, default=0)
    status = db.Column(db.String(20), default='ABERTO')
    hora_abertura = db.Column(db.DateTime, default=datetime.utcnow)
    hora_fechamento = db.Column(db.DateTime)
    operador = db.relationship('Usuario', backref='caixas')

class FormaPagamento(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(50), nullable=False, unique=True)
    ativo = db.Column(db.Boolean, default=True)

class BandeiraCartao(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(50), nullable=False, unique=True)
    ativo = db.Column(db.Boolean, default=True)

class CategoriaDespesa(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    tipo = db.Column(db.String(20), nullable=False)  # FIXA, VARIAVEL, SAIDA
    ativo = db.Column(db.Boolean, default=True)

class Motoboy(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(100), nullable=False)
    taxa_padrao = db.Column(db.Float, default=5.00)
    ativo = db.Column(db.Boolean, default=True)

class Venda(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    caixa_id = db.Column(db.Integer, db.ForeignKey('caixa.id'))
    tipo = db.Column(db.String(20), nullable=False)  # MESA, BALCAO
    numero = db.Column(db.Integer)
    total = db.Column(db.Float, nullable=False)
    emitiu_nota = db.Column(db.Boolean, default=False)
    observacao = db.Column(db.Text)
    data_hora = db.Column(db.DateTime, default=datetime.utcnow)
    caixa = db.relationship('Caixa', backref='vendas')

class PagamentoVenda(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    venda_id = db.Column(db.Integer, db.ForeignKey('venda.id'))
    forma_pagamento_id = db.Column(db.Integer, db.ForeignKey('forma_pagamento.id'))
    bandeira_id = db.Column(db.Integer, db.ForeignKey('bandeira_cartao.id'))
    valor = db.Column(db.Float, nullable=False)
    observacao = db.Column(db.String(200))
    venda = db.relationship('Venda', backref='pagamentos')
    forma_pagamento = db.relationship('FormaPagamento')
    bandeira = db.relationship('BandeiraCartao')

class Delivery(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    caixa_id = db.Column(db.Integer, db.ForeignKey('caixa.id'))
    cliente = db.Column(db.String(100), nullable=False)
    total = db.Column(db.Float, nullable=False)
    taxa_entrega = db.Column(db.Float, default=0)
    motoboy_id = db.Column(db.Integer, db.ForeignKey('motoboy.id'))
    emitiu_nota = db.Column(db.Boolean, default=False)
    observacao = db.Column(db.Text)
    data_hora = db.Column(db.DateTime, default=datetime.utcnow)
    caixa = db.relationship('Caixa', backref='deliveries')
    motoboy = db.relationship('Motoboy')

class PagamentoDelivery(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    delivery_id = db.Column(db.Integer, db.ForeignKey('delivery.id'))
    forma_pagamento_id = db.Column(db.Integer, db.ForeignKey('forma_pagamento.id'))
    bandeira_id = db.Column(db.Integer, db.ForeignKey('bandeira_cartao.id'))
    valor = db.Column(db.Float, nullable=False)
    observacao = db.Column(db.String(200))
    delivery = db.relationship('Delivery', backref='pagamentos')
    forma_pagamento = db.relationship('FormaPagamento')
    bandeira = db.relationship('BandeiraCartao')

class Despesa(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    caixa_id = db.Column(db.Integer, db.ForeignKey('caixa.id'))
    tipo = db.Column(db.String(20), nullable=False)  # FIXA, VARIAVEL, SAIDA
    categoria_id = db.Column(db.Integer, db.ForeignKey('categoria_despesa.id'))
    descricao = db.Column(db.String(200), nullable=False)
    valor = db.Column(db.Float, nullable=False)
    forma_pagamento_id = db.Column(db.Integer, db.ForeignKey('forma_pagamento.id'))
    data_vencimento = db.Column(db.Date)
    status = db.Column(db.String(20), default='PAGO')
    observacao = db.Column(db.Text)
    data_hora = db.Column(db.DateTime, default=datetime.utcnow)
    caixa = db.relationship('Caixa', backref='despesas')
    categoria = db.relationship('CategoriaDespesa')
    forma_pagamento = db.relationship('FormaPagamento')

class Sangria(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    caixa_id = db.Column(db.Integer, db.ForeignKey('caixa.id'))
    valor = db.Column(db.Float, nullable=False)
    motivo = db.Column(db.String(100), nullable=False)
    observacao = db.Column(db.String(200))
    data_hora = db.Column(db.DateTime, default=datetime.utcnow)
    caixa = db.relationship('Caixa', backref='sangrias')

# ==================== MODELO SUPRIMENTO (v3.0) ====================
class Suprimento(db.Model):
    """Suprimentos - Entradas de dinheiro no caixa"""
    __tablename__ = 'suprimento'
    id = db.Column(db.Integer, primary_key=True)
    caixa_id = db.Column(db.Integer, db.ForeignKey('caixa.id'))
    valor = db.Column(db.Float, nullable=False)
    motivo = db.Column(db.String(100), nullable=False)
    observacao = db.Column(db.String(200))
    data_hora = db.Column(db.DateTime, default=datetime.utcnow)
    caixa = db.relationship('Caixa', backref='suprimentos')

class Produto(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    codigo = db.Column(db.String(50), unique=True)
    nome = db.Column(db.String(200), nullable=False)
    categoria = db.Column(db.String(50))
    tamanho = db.Column(db.String(100))
    custo = db.Column(db.Float, default=0)
    preco_venda = db.Column(db.Float, default=0)
    quantidade = db.Column(db.Integer, default=0)
    estoque_minimo = db.Column(db.Integer, default=0)
    estoque_maximo = db.Column(db.Integer, default=0)
    unidade = db.Column(db.String(10), default='UN')
    ativo = db.Column(db.Boolean, default=True)
    data_criacao = db.Column(db.DateTime, default=datetime.utcnow)

class Compra(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    data = db.Column(db.DateTime, default=datetime.utcnow)
    tipo = db.Column(db.String(20), default='BALCAO')
    fornecedor = db.Column(db.String(200))
    total = db.Column(db.Float, default=0)
    observacao = db.Column(db.Text)
    itens = db.relationship('CompraItem', backref='compra', cascade='all, delete-orphan')

class CompraItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    compra_id = db.Column(db.Integer, db.ForeignKey('compra.id'))
    nome = db.Column(db.String(200), nullable=False)
    codigo = db.Column(db.String(50))
    quantidade = db.Column(db.Float, default=0)
    unidade = db.Column(db.String(10), default='UN')
    preco_unitario = db.Column(db.Float, default=0)

class MovimentacaoEstoque(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    produto_id = db.Column(db.Integer, db.ForeignKey('produto.id'))
    tipo = db.Column(db.String(20), nullable=False)  # ENTRADA, SAIDA, AJUSTE
    quantidade = db.Column(db.Integer, nullable=False)
    valor_unitario = db.Column(db.Float, default=0)
    valor_total = db.Column(db.Float, default=0)
    motivo = db.Column(db.String(100))
    observacao = db.Column(db.Text)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'))
    data_hora = db.Column(db.DateTime, default=datetime.utcnow)
    produto = db.relationship('Produto', backref='movimentacoes')
    usuario = db.relationship('Usuario')

# ==================== MODELOS DE LICENÃ‡A ====================

class Licenca(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(100), nullable=False, unique=True)
    chave_ativacao = db.Column(db.String(50), nullable=False, unique=True)
    data_ativacao = db.Column(db.DateTime, default=datetime.utcnow)
    data_expiracao = db.Column(db.DateTime)
    status = db.Column(db.String(20), default='ATIVA')  # ATIVA, EXPIRADA, BLOQUEADA
    max_dispositivos = db.Column(db.Integer, default=2)
    ativo = db.Column(db.Boolean, default=True)

class Dispositivo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    licenca_id = db.Column(db.Integer, db.ForeignKey('licenca.id'))
    nome = db.Column(db.String(100))
    endereco_ip = db.Column(db.String(50))
    mac_address = db.Column(db.String(50))
    user_agent = db.Column(db.String(200))
    data_registro = db.Column(db.DateTime, default=datetime.utcnow)
    ultimo_acesso = db.Column(db.DateTime, default=datetime.utcnow)
    status = db.Column(db.String(20), default='ATIVO')  # ATIVO, BLOQUEADO
    dispositivo_id = db.Column(db.String(100), unique=True)  # Hash Ãºnico do dispositivo
    licenca = db.relationship('Licenca', backref='dispositivos')

class Backup(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nome_arquivo = db.Column(db.String(200))
    tamanho = db.Column(db.Integer)
    data_backup = db.Column(db.DateTime, default=datetime.utcnow)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'))
    observacao = db.Column(db.Text)
    usuario = db.relationship('Usuario')

# ==================== HELPERS - COMPRAS / PRODUTOS ====================

def _local_tag(tag):
    return tag.split('}', 1)[-1] if '}' in tag else tag

def parse_nfe_xml(xml_file):
    """Extrai itens bÃ¡sicos de uma NFe/XML de compra."""
    import xml.etree.ElementTree as ET

    tree = ET.parse(xml_file)
    root = tree.getroot()
    itens = []

    for det in root.iter():
        if _local_tag(det.tag).lower() != 'det':
            continue

        nome = ''
        codigo = ''
        unidade = 'UN'
        quantidade = 0
        preco_unitario = 0

        for node in det.iter():
            tag = _local_tag(node.tag).lower()
            text = (node.text or '').strip()
            if tag == 'xprod':
                nome = text
            elif tag == 'cprod':
                codigo = text
            elif tag == 'ucom':
                unidade = text or 'UN'
            elif tag == 'qcom':
                try:
                    quantidade = float(text.replace(',', '.'))
                except Exception:
                    quantidade = 0
            elif tag == 'vuncom':
                try:
                    preco_unitario = float(text.replace(',', '.'))
                except Exception:
                    preco_unitario = 0

        if not nome:
            continue

        itens.append({
            'nome': nome,
            'codigo': codigo,
            'quantidade': quantidade,
            'unidade': unidade,
            'preco_unitario': preco_unitario
        })

    return itens

def _find_or_create_produto(nome, codigo=None, preco_unitario=0, quantidade=0, unidade='UN'):
    produto = None
    if codigo:
        produto = Produto.query.filter_by(codigo=codigo).first()
    if not produto:
        produto = Produto.query.filter_by(nome=nome).first()

    if produto:
        produto.quantidade = (produto.quantidade or 0) + float(quantidade or 0)
        if preco_unitario:
            produto.custo = float(preco_unitario)
        if unidade:
            produto.unidade = unidade
        return produto

    produto = Produto(
        codigo=codigo or f'PROD{Produto.query.count() + 1:03d}',
        nome=nome,
        custo=float(preco_unitario or 0),
        quantidade=int(float(quantidade or 0)),
        unidade=unidade or 'UN'
    )
    db.session.add(produto)
    return produto

def _parse_excel_date(value):
    if value is None or value == '':
        return datetime.utcnow()
    if isinstance(value, datetime):
        return value
    if hasattr(value, 'year') and hasattr(value, 'month') and hasattr(value, 'day'):
        return datetime(value.year, value.month, value.day)
    text = str(value).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(text, fmt)
        except Exception:
            pass
    return datetime.utcnow()

def _normalizar_turno(turno):
    """Normaliza o texto do turno para comparação segura."""
    if not turno:
        return ''
    texto = unicodedata.normalize('NFKD', str(turno))
    texto = texto.encode('ascii', 'ignore').decode('ascii')
    return texto.upper().strip()

def _turno_canonico(turno):
    texto = _normalizar_turno(turno)
    if 'MANH' in texto:
        return 'MANHÃ'
    if 'TAR' in texto:
        return 'TARDE'
    if 'NOI' in texto:
        return 'NOITE'
    return texto or 'MANHÃ'

def _buscar_caixa_por_data_turno(data_ref, turno_ref=None):
    """Localiza um caixa pela data e turno, aceitando variações de acentuação."""
    if isinstance(data_ref, datetime):
        data_ref = data_ref.date()
    if not data_ref:
        return None

    caixas = Caixa.query.filter_by(data=data_ref).order_by(Caixa.id.desc()).all()
    if not caixas:
        return None

    if turno_ref:
        turno_ref_norm = _normalizar_turno(turno_ref)
        for caixa in caixas:
            if _normalizar_turno(caixa.turno) == turno_ref_norm:
                return caixa

    return caixas[0]

def _buscar_caixas_por_data_turno(data_ref, turno_ref=None):
    """Retorna todos os caixas de uma data/turno, com tolerância a acentos."""
    if isinstance(data_ref, datetime):
        data_ref = data_ref.date()
    if not data_ref:
        return []

    caixas = Caixa.query.filter_by(data=data_ref).order_by(Caixa.id.desc()).all()
    if not caixas:
        return []

    if turno_ref:
        turno_ref_norm = _normalizar_turno(turno_ref)
        filtrados = [caixa for caixa in caixas if _normalizar_turno(caixa.turno) == turno_ref_norm]
        if filtrados:
            return filtrados

    return caixas

def _intervalo_periodo(data_ref, periodo):
    """Calcula intervalo (início, fim) inclusive para dia/semana/mês."""
    if isinstance(data_ref, datetime):
        data_ref = data_ref.date()
    if not data_ref:
        hoje = datetime.utcnow().date()
        data_ref = hoje

    periodo = (periodo or 'dia').lower().strip()
    if periodo not in ('dia', 'semana', 'mes'):
        periodo = 'dia'

    if periodo == 'dia':
        return data_ref, data_ref

    if periodo == 'semana':
        # Últimos 7 dias (inclui a data de referência)
        inicio = data_ref - timedelta(days=6)
        fim = data_ref
        return inicio, fim

    # mes
    inicio = data_ref.replace(day=1)
    if inicio.month == 12:
        proximo = inicio.replace(year=inicio.year + 1, month=1, day=1)
    else:
        proximo = inicio.replace(month=inicio.month + 1, day=1)
    fim = proximo - timedelta(days=1)
    return inicio, fim

def _buscar_caixas_por_intervalo(data_inicio, data_fim, turno_ref=None):
    """Retorna caixas por intervalo de datas (inclusive) e opcionalmente por turno."""
    if isinstance(data_inicio, datetime):
        data_inicio = data_inicio.date()
    if isinstance(data_fim, datetime):
        data_fim = data_fim.date()

    if not data_inicio or not data_fim:
        return []
    if data_inicio > data_fim:
        data_inicio, data_fim = data_fim, data_inicio

    caixas = (
        Caixa.query
        .filter(Caixa.data >= data_inicio, Caixa.data <= data_fim)
        .order_by(Caixa.data.desc(), Caixa.id.desc())
        .all()
    )
    if not caixas:
        return []

    turno_ref = (turno_ref or '').strip()
    if not turno_ref or _normalizar_turno(turno_ref) in ('TODOS', 'TODAS'):
        return caixas

    turno_ref_norm = _normalizar_turno(turno_ref)
    filtrados = [caixa for caixa in caixas if _normalizar_turno(caixa.turno) == turno_ref_norm]
    return filtrados or caixas

def _buscar_caixas_por_periodo(data_ref, periodo, turno_ref=None):
    """Retorna caixas por período (dia/semana/mês) e opcionalmente por turno."""
    inicio, fim = _intervalo_periodo(data_ref, periodo)
    caixas = (
        Caixa.query
        .filter(Caixa.data >= inicio, Caixa.data <= fim)
        .order_by(Caixa.data.desc(), Caixa.id.desc())
        .all()
    )
    if not caixas:
        return []

    turno_ref = (turno_ref or '').strip()
    if not turno_ref or _normalizar_turno(turno_ref) in ('TODOS', 'TODAS'):
        return caixas

    turno_ref_norm = _normalizar_turno(turno_ref)
    filtrados = [caixa for caixa in caixas if _normalizar_turno(caixa.turno) == turno_ref_norm]
    return filtrados or caixas

def _resolver_caixa_visualizacao():
    """Define filtros (data/turno/período) para as telas operacionais."""
    caixa_sessao = db.session.get(Caixa, session.get('caixa_id')) if session.get('caixa_id') else None
    data_ref = request.args.get('data')
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    turno_ref = request.args.get('turno')
    periodo_ref = (request.args.get('periodo') or 'dia').lower().strip()
    if periodo_ref not in ('dia', 'semana', 'mes'):
        periodo_ref = 'dia'

    data_inicio_dt = None
    data_fim_dt = None
    if data_inicio or data_fim:
        try:
            data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d').date() if data_inicio else None
        except Exception:
            data_inicio_dt = None
        try:
            data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d').date() if data_fim else None
        except Exception:
            data_fim_dt = None

        # Defaults when only one side is provided
        if data_inicio_dt and not data_fim_dt:
            data_fim_dt = data_inicio_dt
        if data_fim_dt and not data_inicio_dt:
            data_inicio_dt = data_fim_dt

    if data_ref:
        try:
            data_ref = datetime.strptime(data_ref, '%Y-%m-%d').date()
        except Exception:
            data_ref = caixa_sessao.data if caixa_sessao else datetime.utcnow().date()
    else:
        data_ref = caixa_sessao.data if caixa_sessao else datetime.utcnow().date()

    if not turno_ref:
        if periodo_ref in ('semana', 'mes'):
            turno_ref = 'TODOS'
        else:
            turno_ref = caixa_sessao.turno if caixa_sessao else session.get('turno')

    caixa_filtrada = _buscar_caixa_por_data_turno(data_ref, turno_ref)
    if caixa_filtrada:
        turno_out = 'TODOS' if _normalizar_turno(turno_ref) in ('TODOS', 'TODAS') else _turno_canonico(caixa_filtrada.turno)
        return caixa_filtrada, data_ref, turno_out, periodo_ref, data_inicio_dt, data_fim_dt

    turno_out = 'TODOS' if _normalizar_turno(turno_ref) in ('TODOS', 'TODAS') else _turno_canonico(turno_ref or (caixa_sessao.turno if caixa_sessao else 'MANHÃ'))
    return caixa_sessao, data_ref, turno_out, periodo_ref, data_inicio_dt, data_fim_dt

def _calcular_totais_caixas(caixas):
    """Agrega os totais de uma lista de caixas."""
    totais = {
        'vendas_loja': 0,
        'vendas_delivery': 0,
        'total_vendas': 0,
        'dinheiro': 0,
        'credito': 0,
        'debito': 0,
        'pix': 0,
        'online': 0,
        'notas_fiscais': 0,
        'despesas': 0,
        'sangrias': 0,
        'saldo_atual': 0,
    }

    for caixa in caixas or []:
        totais_caixa = calcular_totais_caixa(caixa)
        for key in totais:
            totais[key] += totais_caixa.get(key, 0)

    return totais

def _calcular_totais_delivery_caixas(caixas):
    """Agrega totais de delivery para uma lista de caixas."""
    totais = {
        'total_delivery': 0,
        'total_taxas': 0,
        'quantidade_pedidos': 0,
        'motoboys': {}
    }

    for caixa in caixas or []:
        totais_caixa = calcular_totais_delivery(caixa)
        totais['total_delivery'] += totais_caixa.get('total_delivery', 0)
        totais['total_taxas'] += totais_caixa.get('total_taxas', 0)
        totais['quantidade_pedidos'] += totais_caixa.get('quantidade_pedidos', 0)
        for nome, valor in totais_caixa.get('motoboys', {}).items():
            totais['motoboys'][nome] = totais['motoboys'].get(nome, 0) + valor

    return totais

def _obter_caixa_para_data_importacao(data_ref, turno_preferencial=None):
    """Busca um caixa existente pela data da planilha ou cria um caixa novo para importação."""
    if isinstance(data_ref, datetime):
        data_ref = data_ref.date()
    elif hasattr(data_ref, 'date') and not isinstance(data_ref, date):
        try:
            data_ref = data_ref.date()
        except Exception:
            data_ref = datetime.utcnow().date()

    if not data_ref:
        data_ref = datetime.utcnow().date()

    if turno_preferencial:
        caixa = (
            Caixa.query
            .filter_by(data=data_ref, turno=turno_preferencial)
            .order_by(Caixa.id.desc())
            .first()
        )
        if caixa:
            return caixa

    caixa = (
        Caixa.query
        .filter_by(data=data_ref)
        .order_by(Caixa.id.desc())
        .first()
    )
    if caixa:
        return caixa

    turno = (turno_preferencial or session.get('turno') or 'MANHÃ').strip().upper()
    if turno not in ['MANHÃ', 'TARDE', 'NOITE']:
        turno = 'MANHÃ'

    caixa = Caixa(
        data=data_ref,
        turno=turno,
        operador_id=session.get('user_id'),
        saldo_inicial=0,
        status='ABERTO'
    )
    db.session.add(caixa)
    db.session.flush()
    return caixa

def _lookup_or_create_forma_pagamento(nome):
    if not nome:
        return None
    from sqlalchemy import func
    forma = FormaPagamento.query.filter(func.lower(FormaPagamento.nome) == str(nome).strip().lower()).first()
    if not forma:
        forma = FormaPagamento(nome=str(nome).strip(), ativo=True)
        db.session.add(forma)
        db.session.flush()
    return forma

def _lookup_or_create_bandeira(nome):
    if not nome:
        return None
    from sqlalchemy import func
    bandeira = BandeiraCartao.query.filter(func.lower(BandeiraCartao.nome) == str(nome).strip().lower()).first()
    if not bandeira:
        bandeira = BandeiraCartao(nome=str(nome).strip(), ativo=True)
        db.session.add(bandeira)
        db.session.flush()
    return bandeira

def _lookup_or_create_motoboy(nome):
    if not nome:
        return None
    from sqlalchemy import func
    motoboy = Motoboy.query.filter(func.lower(Motoboy.nome) == str(nome).strip().lower()).first()
    if not motoboy:
        motoboy = Motoboy(nome=str(nome).strip(), taxa_padrao=5.00, ativo=True)
        db.session.add(motoboy)
        db.session.flush()
    return motoboy

# ==================== DECORATORS ====================

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def dashboard_required(f):
    """Requer permissÃ£o para acessar dashboard"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        usuario = db.session.get(Usuario, session['user_id'])
        # Verifica se o usuário tem permissão para acessar o dashboard
        if not usuario or not getattr(usuario, 'acesso_dashboard', False):
            flash('Acesso negado ao dashboard.', 'danger')
            return redirect(url_for('vendas'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        usuario = db.session.get(Usuario, session['user_id'])
        if not usuario or usuario.perfil not in ['ADMIN', 'MASTER']:
            flash('Acesso negado. Voce nao tem permissao para acessar esta area.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function


def admin_master_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        usuario = db.session.get(Usuario, session['user_id'])
        if not usuario or usuario.perfil != 'MASTER':
            flash('Acesso negado. Apenas o ADMIN MASTER pode acessar esta area.', 'danger')
            return redirect(url_for('vendas'))
        return f(*args, **kwargs)
    return decorated_function


# ==================== ROUTES - AUTH ====================

@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('vendas'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        try:
            nome = request.form.get('operador') or request.form.get('username')
            senha = request.form.get('senha') or request.form.get('password')
            acao = request.form.get('acao', 'novo')
            usuario = Usuario.query.filter_by(nome=nome, ativo=True).first()

            if usuario and check_password_hash(usuario.senha, senha):
                if usuario.perfil == 'MASTER':
                    usuario.acesso_dashboard = True
                    usuario.acesso_configuracoes = True
                    usuario.acesso_relatorios = True
                    db.session.commit()

                # ACESSAR CAIXA EXISTENTE
                if acao == 'acessar':
                    caixa_id_raw = request.form.get('caixa_id')
                    if not caixa_id_raw:
                        flash('Selecione um caixa para acessar.', 'danger')
                        return redirect(url_for('login'))

                    caixa = db.session.get(Caixa, int(caixa_id_raw))

                    if not caixa:
                        flash('Caixa não encontrado!', 'danger')
                        return redirect(url_for('login'))

                    # Restaurar sessão com TODAS as informações
                    session['user_id'] = usuario.id
                    session['user_nome'] = usuario.nome
                    session['caixa_id'] = caixa.id
                    session['turno'] = caixa.turno
                    session['data'] = caixa.data.strftime('%Y-%m-%d')
                    session['saldo_inicial'] = caixa.saldo_inicial
                    session['hora_abertura'] = caixa.hora_abertura.strftime('%H:%M:%S')

                    # Calcular totais atuais
                    totais = calcular_totais_caixa(caixa)
                    session['total_vendas_atual'] = totais['total_vendas']
                    session['total_despesas_atual'] = totais['despesas']
                    session['saldo_atual'] = totais['saldo_atual']

                    flash(f'✅ Bem-vindo de volta, {usuario.nome}! Continuando caixa #{caixa.id} - {caixa.turno}', 'success')
                    return redirect(url_for('vendas'))

                # ABRIR NOVO CAIXA
                else:
                    data = request.form.get('data') or datetime.now().strftime('%Y-%m-%d')
                    turno = (request.form.get('turno') or 'MANHÃ').strip().upper()
                    saldo_inicial = parse_moeda(request.form.get('saldo_inicial', 100))

                    data_obj = datetime.strptime(data, '%Y-%m-%d').date()

                    caixa_aberto = Caixa.query.filter_by(data=data_obj, turno=turno, status='ABERTO').first()
                    caixa_fechado = Caixa.query.filter_by(data=data_obj, turno=turno, status='FECHADO').first()
                if usuario.perfil in ['ADMIN', 'MASTER']:
                    if caixa_aberto:
                        caixa_para_usar = caixa_aberto
                        flash(f'Admin assumindo caixa ABERTO - {turno}', 'info')
                    elif caixa_fechado:
                        caixa_fechado.status = 'ABERTO'
                        caixa_fechado.hora_fechamento = None
                        db.session.commit()
                        caixa_para_usar = caixa_fechado
                        flash(f'Caixa FECHADO reaberto pelo Admin', 'warning')
                    else:
                        caixa_para_usar = Caixa(
                            data=data_obj, turno=turno, operador_id=usuario.id,
                            saldo_inicial=saldo_inicial, status='ABERTO'
                        )
                        db.session.add(caixa_para_usar)
                        db.session.commit()
                        flash('OK: Novo caixa aberto com sucesso!', 'success')
                    
                    session['user_id'] = usuario.id
                    session['user_nome'] = usuario.nome
                    session['caixa_id'] = caixa_para_usar.id
                    session['turno'] = turno
                    session['data'] = data
                    session['saldo_inicial'] = caixa_para_usar.saldo_inicial
                    session['hora_abertura'] = caixa_para_usar.hora_abertura.strftime('%H:%M:%S')
                    return redirect(url_for('vendas'))
                else:
                    if caixa_aberto:
                        # Operador pode ENTRAR no caixa aberto
                        session['user_id'] = usuario.id
                        session['user_nome'] = usuario.nome
                        session['caixa_id'] = caixa_aberto.id
                        session['turno'] = caixa_aberto.turno
                        session['data'] = caixa_aberto.data.strftime('%Y-%m-%d')
                        session['saldo_inicial'] = caixa_aberto.saldo_inicial
                        session['hora_abertura'] = caixa_aberto.hora_abertura.strftime('%H:%M:%S')

                        flash(f'OK: Bem-vindo, {usuario.nome}! Caixa do turno {turno} acessado.', 'success')
                        return redirect(url_for('vendas'))

                    if caixa_fechado:
                        # Permitir operador acessar caixa fechado para visualizaÃ§Ã£o
                        session['user_id'] = usuario.id
                        session['user_nome'] = usuario.nome
                        session['caixa_id'] = caixa_fechado.id
                        session['turno'] = caixa_fechado.turno
                        session['data'] = caixa_fechado.data.strftime('%Y-%m-%d')
                        session['saldo_inicial'] = caixa_fechado.saldo_inicial
                        session['hora_abertura'] = caixa_fechado.hora_abertura.strftime('%H:%M:%S')
                        
                        flash(f'Aviso: Acessando caixa FECHADO de {turno} (Modo Visualização).', 'warning')
                        return redirect(url_for('vendas'))
                    
                    novo_caixa = Caixa(
                        data=data_obj, turno=turno, operador_id=usuario.id,
                        saldo_inicial=saldo_inicial, status='ABERTO'
                    )
                    db.session.add(novo_caixa)
                    db.session.commit()
                    
                    session['user_id'] = usuario.id
                    session['user_nome'] = usuario.nome
                    session['caixa_id'] = novo_caixa.id
                    session['turno'] = turno
                    session['data'] = data
                    session['saldo_inicial'] = novo_caixa.saldo_inicial
                    session['hora_abertura'] = novo_caixa.hora_abertura.strftime('%H:%M:%S')
                    flash('OK: Caixa aberto com sucesso!', 'success')
                    return redirect(url_for('vendas'))
            else:
                flash('❌ Credenciais inválidas!', 'danger')
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao acessar o sistema: {str(e)}', 'danger')
    
    try:
        usuarios = Usuario.query.filter_by(ativo=True).order_by(Usuario.nome).all()
    except ProgrammingError:
        db.session.rollback()
        _ensure_db_ready()
        usuarios = Usuario.query.filter_by(ativo=True).order_by(Usuario.nome).all()
    hoje = datetime.now().date()
    caixas_abertos_hoje = Caixa.query.filter_by(data=hoje, status='ABERTO').order_by(Caixa.turno).all()
    
    return render_template('login.html', usuarios=usuarios, caixas_abertos=caixas_abertos_hoje)

@app.route('/logout')
def logout():
    session.clear()
    flash('Logout realizado com sucesso!', 'info')
    return redirect(url_for('login'))


# ==================== ROUTES - LICENCIAMENTO ====================

def _gerar_chave_ativacao():
    return '-'.join(''.join(random.choices(string.ascii_uppercase + string.digits, k=4)) for _ in range(4))

def _device_fingerprint():
    ip = request.remote_addr or '0.0.0.0'
    ua = request.user_agent.string or 'unknown'
    raw = f"{ip}|{ua}"
    return hashlib.md5(raw.encode('utf-8')).hexdigest(), ip, ua

@app.route('/ativacao', methods=['GET', 'POST'])
def ativacao():
    if request.method == 'POST':
        email = request.form.get('email')
        chave = request.form.get('chave')
        if not email or not chave:
            flash('Informe e-mail e chave para ativar.', 'warning')
            return redirect(url_for('ativacao'))

        licenca = Licenca.query.filter_by(email=email, chave_ativacao=chave).first()
        if not licenca:
            licenca = Licenca(
                email=email,
                chave_ativacao=chave,
                data_ativacao=datetime.utcnow(),
                data_expiracao=datetime.utcnow() + timedelta(days=365),
                status='ATIVA',
                ativo=True
            )
            db.session.add(licenca)
            db.session.commit()

        flash('Sistema ativado com sucesso!', 'success')
        return redirect(url_for('login'))

    return render_template('ativacao.html')

@app.route('/licenciamento')
@login_required
@admin_master_required
def licenciamento():
    licenca = Licenca.query.first()
    backups = Backup.query.order_by(Backup.data_backup.desc()).all()
    dispositivos = licenca.dispositivos if licenca else []

    dispositivo_id, ip, ua = _device_fingerprint()
    if licenca:
        dispositivo = Dispositivo.query.filter_by(dispositivo_id=dispositivo_id, licenca_id=licenca.id).first()
        if not dispositivo:
            dispositivo = Dispositivo(
                licenca_id=licenca.id,
                nome=request.user_agent.browser or 'Dispositivo',
                endereco_ip=ip,
                user_agent=ua,
                dispositivo_id=dispositivo_id,
                status='ATIVO'
            )
            db.session.add(dispositivo)
        dispositivo.ultimo_acesso = datetime.utcnow()
        db.session.commit()

    return render_template(
        'licenciamento.html',
        licenca=licenca,
        dispositivos=dispositivos,
        backups=backups,
        dispositivo_id=dispositivo_id
    )

@app.route('/licenca/registrar-dispositivo', methods=['POST'])
@login_required
@admin_master_required
def registrar_dispositivo():
    licenca = Licenca.query.first()
    if not licenca or not licenca.ativo:
        return jsonify({'status': 'no_license', 'error': 'LicenÃ§a nÃ£o ativa'}), 400

    dispositivo_id, ip, ua = _device_fingerprint()
    dispositivo = Dispositivo.query.filter_by(dispositivo_id=dispositivo_id, licenca_id=licenca.id).first()

    if not dispositivo:
        if len(licenca.dispositivos) >= licenca.max_dispositivos:
            return jsonify({'status': 'limit_reached', 'error': 'Limite de dispositivos atingido'}), 400
        dispositivo = Dispositivo(
            licenca_id=licenca.id,
            nome=request.user_agent.browser or 'Dispositivo',
            endereco_ip=ip,
            user_agent=ua,
            dispositivo_id=dispositivo_id,
            status='ATIVO'
        )
        db.session.add(dispositivo)

    dispositivo.ultimo_acesso = datetime.utcnow()
    db.session.commit()
    return jsonify({'status': 'ok'})

@app.route('/licenca/gerar-nova-chave', methods=['POST'])
@login_required
@admin_master_required
def gerar_nova_chave():
    licenca = Licenca.query.first()
    if not licenca:
        flash('Nenhuma licenÃ§a encontrada.', 'warning')
        return redirect(url_for('licenciamento'))
    licenca.chave_ativacao = _gerar_chave_ativacao()
    licenca.status = 'ATIVA'
    licenca.ativo = True
    db.session.commit()
    flash('Nova chave gerada com sucesso!', 'success')
    return redirect(url_for('licenciamento'))

@app.route('/licenca/bloquear-todos-dispositivos', methods=['POST'])
@login_required
@admin_master_required
def bloquear_todos_dispositivos():
    licenca = Licenca.query.first()
    if licenca:
        Dispositivo.query.filter_by(licenca_id=licenca.id).update({'status': 'BLOQUEADO'})
        db.session.commit()
    flash('Todos os dispositivos foram bloqueados.', 'warning')
    return redirect(url_for('licenciamento'))

@app.route('/licenca/bloquear-dispositivo/<int:dispositivo_id>', methods=['POST'])
@login_required
@admin_master_required
def bloquear_dispositivo(dispositivo_id):
    dispositivo = db.session.get(Dispositivo, dispositivo_id)
    if dispositivo:
        dispositivo.status = 'BLOQUEADO'
        db.session.commit()
        flash('Dispositivo bloqueado.', 'warning')
    return redirect(url_for('licenciamento'))

@app.route('/licenca/desbloquear-dispositivo/<int:dispositivo_id>', methods=['POST'])
@login_required
@admin_master_required
def desbloquear_dispositivo(dispositivo_id):
    dispositivo = db.session.get(Dispositivo, dispositivo_id)
    if dispositivo:
        dispositivo.status = 'ATIVO'
        db.session.commit()
        flash('Dispositivo desbloqueado.', 'success')
    return redirect(url_for('licenciamento'))

@app.route('/licenca/excluir-dispositivo/<int:dispositivo_id>', methods=['POST'])
@login_required
@admin_master_required
def excluir_dispositivo(dispositivo_id):
    dispositivo = db.session.get(Dispositivo, dispositivo_id)
    if dispositivo:
        db.session.delete(dispositivo)
        db.session.commit()
        flash('Dispositivo excluÃ­do.', 'success')
    return redirect(url_for('licenciamento'))

@app.route('/licenca/upload-backup', methods=['POST'])
@login_required
@admin_master_required
def upload_backup():
    arquivo = request.files.get('backup')
    observacao = request.form.get('observacao', '')
    if not arquivo or arquivo.filename == '':
        flash('Selecione um arquivo de backup.', 'warning')
        return redirect(url_for('licenciamento'))
    if not arquivo.filename.lower().endswith('.db'):
        flash('Apenas arquivos .db sÃ£o permitidos.', 'danger')
        return redirect(url_for('licenciamento'))

    pasta = os.path.join(basedir, 'backups')
    os.makedirs(pasta, exist_ok=True)
    nome_arquivo = secure_filename(arquivo.filename)
    caminho = os.path.join(pasta, nome_arquivo)
    arquivo.save(caminho)

    backup = Backup(
        nome_arquivo=nome_arquivo,
        tamanho=os.path.getsize(caminho),
        usuario_id=session.get('user_id'),
        observacao=observacao
    )
    db.session.add(backup)
    db.session.commit()
    flash('Backup enviado com sucesso.', 'success')
    return redirect(url_for('licenciamento'))

@app.route('/licenca/download-backup/<int:backup_id>')
@login_required
@admin_master_required
def download_backup(backup_id):
    backup = db.session.get(Backup, backup_id)
    if not backup:
        flash('Backup nÃ£o encontrado.', 'warning')
        return redirect(url_for('licenciamento'))
    pasta = os.path.join(basedir, 'backups')
    return send_from_directory(pasta, backup.nome_arquivo, as_attachment=True)

@app.route('/licenca/excluir-backup/<int:backup_id>', methods=['POST'])
@login_required
@admin_master_required
def excluir_backup(backup_id):
    backup = db.session.get(Backup, backup_id)
    if backup:
        pasta = os.path.join(basedir, 'backups')
        caminho = os.path.join(pasta, backup.nome_arquivo)
        if os.path.exists(caminho):
            os.remove(caminho)
        db.session.delete(backup)
        db.session.commit()
        flash('Backup excluÃ­do.', 'success')
    return redirect(url_for('licenciamento'))

# ==================== ROUTES - VENDAS ====================

@app.route('/vendas')
@login_required
def vendas():
    try:
        caixa, data_filtro, turno_filtro, periodo_filtro, data_inicio_filtro, data_fim_filtro = _resolver_caixa_visualizacao()
        if data_inicio_filtro and data_fim_filtro:
            caixas_filtro = _buscar_caixas_por_intervalo(data_inicio_filtro, data_fim_filtro, turno_filtro)
        else:
            caixas_filtro = _buscar_caixas_por_periodo(data_filtro, periodo_filtro, turno_filtro)
        if not caixas_filtro:
            caixas_filtro = [caixa] if caixa else []
        if not caixas_filtro:
            flash('Abra ou acesse um caixa para continuar.', 'warning')
            return redirect(url_for('login'))

        caixa_ids = [c.id for c in caixas_filtro]
        vendas = Venda.query.filter(Venda.caixa_id.in_(caixa_ids)).order_by(Venda.data_hora.desc()).all()
        formas_pagamento = FormaPagamento.query.filter_by(ativo=True).all()
        bandeiras = BandeiraCartao.query.filter_by(ativo=True).all()

        if data_inicio_filtro and data_fim_filtro:
            inicio, fim = data_inicio_filtro, data_fim_filtro
        else:
            inicio, fim = _intervalo_periodo(data_filtro, periodo_filtro)
        turnos_raw = (
            Caixa.query.filter(Caixa.data >= inicio, Caixa.data <= fim)
            .with_entities(Caixa.turno)
            .all()
        )
        turnos_set = {_turno_canonico(t[0]) for t in (turnos_raw or []) if t and t[0]}
        ordem = ['MANHÃ', 'TARDE', 'NOITE']
        turnos_disponiveis = ['TODOS'] + [t for t in ordem if t in turnos_set] + sorted([t for t in turnos_set if t not in ordem])

        totais = _calcular_totais_caixas(caixas_filtro)
        totais_delivery = _calcular_totais_delivery_caixas(caixas_filtro)

        # Métricas para o template
        try:
            metricas_avancadas = calcular_metricas_avancadas(caixas_filtro)
        except Exception:
            metricas_avancadas = {}

        try:
            metricas = calcular_metricas_dashboard(caixas_filtro)
        except Exception:
            metricas = {}

        return render_template('vendas.html', 
                             vendas=vendas, 
                             formas_pagamento=formas_pagamento,
                             bandeiras=bandeiras,
                             totais=totais,
                             totais_delivery=totais_delivery,
                             caixa=caixa or caixas_filtro[0],
                             data_filtro=data_filtro,
                             data_inicio_filtro=data_inicio_filtro,
                             data_fim_filtro=data_fim_filtro,
                             turno_filtro=turno_filtro,
                             periodo_filtro=periodo_filtro,
                             turnos_disponiveis=turnos_disponiveis,
                             metricas_avancadas=metricas_avancadas,
                             metricas=metricas)
    except Exception as e:
        db.session.rollback()
        import traceback
        tb = traceback.format_exc()
        print(tb)
        return f"<h1>Erro em /vendas</h1><pre>{tb}</pre>", 500

@app.route('/vendas/nova', methods=['POST'])
@login_required
def nova_venda():
    try:
        tipo = request.form.get('tipo')
        numero = int(request.form.get('numero', 1))
        total = parse_moeda(request.form.get('total'))
        emitiu_nota = request.form.get('emitiu_nota') == 'on'
        observacao = request.form.get('observacao', '')
        
        # Criar venda
        venda = Venda(
            caixa_id=session['caixa_id'],
            tipo=tipo,
            numero=numero,
            total=total,
            emitiu_nota=emitiu_nota,
            observacao=observacao
        )
        db.session.add(venda)
        db.session.flush()
        
        # Adicionar pagamentos
        formas = request.form.getlist('forma_pagamento[]')
        valores = request.form.getlist('valor_pagamento[]')
        bandeiras_ids = request.form.getlist('bandeira[]')
        obs_pagamentos = request.form.getlist('obs_pagamento[]')
        
        total_pago = 0
        for i, forma_id in enumerate(formas):
            if forma_id and valores[i]:
                valor = parse_moeda(valores[i])
                if valor > 0:
                    pagamento = PagamentoVenda(
                        venda_id=venda.id,
                        forma_pagamento_id=int(forma_id),
                        valor=valor,
                        bandeira_id=int(bandeiras_ids[i]) if bandeiras_ids[i] and bandeiras_ids[i] != '' else None,
                        observacao=obs_pagamentos[i] if i < len(obs_pagamentos) else ''
                    )
                    db.session.add(pagamento)
                    total_pago += valor
        
        if abs(total_pago - total) > 0.01:
            db.session.rollback()
            flash('O total dos pagamentos nÃ£o corresponde ao valor da venda!', 'danger')
            return redirect(url_for('vendas'))
        
        db.session.commit()
        flash('Venda registrada com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao registrar venda: {str(e)}', 'danger')
    
    return redirect(url_for('vendas'))

# ==================== ROUTES - DELIVERY ====================

@app.route('/delivery')
@login_required
def delivery():
    try:
        caixa, data_filtro, turno_filtro, periodo_filtro, data_inicio_filtro, data_fim_filtro = _resolver_caixa_visualizacao()
        if data_inicio_filtro and data_fim_filtro:
            caixas_filtro = _buscar_caixas_por_intervalo(data_inicio_filtro, data_fim_filtro, turno_filtro)
        else:
            caixas_filtro = _buscar_caixas_por_periodo(data_filtro, periodo_filtro, turno_filtro)
        if not caixas_filtro:
            caixas_filtro = [caixa] if caixa else []
        if not caixas_filtro:
            flash('Abra ou acesse um caixa para continuar.', 'warning')
            return redirect(url_for('login'))

        caixa_ids = [c.id for c in caixas_filtro]
        deliveries = Delivery.query.filter(Delivery.caixa_id.in_(caixa_ids)).order_by(Delivery.data_hora.desc()).all()
        formas_pagamento = FormaPagamento.query.filter_by(ativo=True).all()
        motoboys = Motoboy.query.filter_by(ativo=True).all()

        if data_inicio_filtro and data_fim_filtro:
            inicio, fim = data_inicio_filtro, data_fim_filtro
        else:
            inicio, fim = _intervalo_periodo(data_filtro, periodo_filtro)
        turnos_raw = (
            Caixa.query.filter(Caixa.data >= inicio, Caixa.data <= fim)
            .with_entities(Caixa.turno)
            .all()
        )
        turnos_set = {_turno_canonico(t[0]) for t in (turnos_raw or []) if t and t[0]}
        ordem = ['MANHÃ', 'TARDE', 'NOITE']
        turnos_disponiveis = ['TODOS'] + [t for t in ordem if t in turnos_set] + sorted([t for t in turnos_set if t not in ordem])

        totais = _calcular_totais_caixas(caixas_filtro)
        totais_delivery = _calcular_totais_delivery_caixas(caixas_filtro)

        bandeiras = BandeiraCartao.query.filter_by(ativo=True).all()
        return render_template('delivery.html', bandeiras=bandeiras,
                             deliveries=deliveries,
                             formas_pagamento=formas_pagamento,
                              motoboys=motoboys,
                              totais=totais,
                              totais_delivery=totais_delivery,
                              caixa=caixa or caixas_filtro[0],
                              data_filtro=data_filtro,
                              data_inicio_filtro=data_inicio_filtro,
                              data_fim_filtro=data_fim_filtro,
                              turno_filtro=turno_filtro,
                              periodo_filtro=periodo_filtro,
                              turnos_disponiveis=turnos_disponiveis)
    except Exception as e:
        db.session.rollback()
        import traceback
        tb = traceback.format_exc()
        print(tb)
        return f"<h1>Erro em /delivery</h1><pre>{tb}</pre>", 500

@app.route('/delivery/novo', methods=['POST'])
@login_required
def novo_delivery():
    try:
        cliente = request.form.get('cliente')
        total = parse_moeda(request.form.get('total'))
        taxa_entrega = parse_moeda(request.form.get('taxa_entrega', 0))
        motoboy_id = request.form.get('motoboy_id')
        emitiu_nota = request.form.get('emitiu_nota') == 'on'
        observacao = request.form.get('observacao', '')
        
        delivery = Delivery(
            caixa_id=session['caixa_id'],
            cliente=cliente,
            total=total,
            taxa_entrega=taxa_entrega,
            motoboy_id=int(motoboy_id) if motoboy_id else None,
            emitiu_nota=emitiu_nota,
            observacao=observacao
        )
        db.session.add(delivery)
        db.session.flush()
        
        # Adicionar pagamentos
        formas = request.form.getlist('forma_pagamento_delivery[]')
        valores = request.form.getlist('valor_pagamento_delivery[]')
        obs_pagamentos = request.form.getlist('obs_pagamento_delivery[]')
        
        total_com_taxa = total + taxa_entrega
        total_pago = 0
        
        for i, forma_id in enumerate(formas):
            if forma_id and valores[i]:
                valor = parse_moeda(valores[i])
                if valor > 0:
                    pagamento = PagamentoDelivery(
                        delivery_id=delivery.id,
                        forma_pagamento_id=int(forma_id),
                        valor=valor,
                        observacao=obs_pagamentos[i] if i < len(obs_pagamentos) else ''
                    )
                    db.session.add(pagamento)
                    total_pago += valor
        
        if abs(total_pago - total_com_taxa) > 0.01:
            db.session.rollback()
            flash('O total dos pagamentos nÃ£o corresponde ao valor total (pedido + taxa)!', 'danger')
            return redirect(url_for('delivery'))
        
        db.session.commit()
        flash('Delivery registrado com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao registrar delivery: {str(e)}', 'danger')
    
    return redirect(url_for('delivery'))

@app.route('/vendas/modelo-planilha')
@login_required
def vendas_modelo_planilha():
    from io import BytesIO
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Vendas'

    headers = ['DATA', 'TIPO', 'N°', 'TOTAL DE MESA', 'PAGAMENTO', 'FORMA', 'BANDEIRA', 'NOTA FISCAL', 'OBSERVAÇÃO']
    ws.append(headers)
    ws.append(['07/05/2026', 'MESA', 5, 84.70, 84.70, 'DINHEIRO', '', 'NAO', 'Exemplo'])

    fill = PatternFill('solid', fgColor='1F3B73')
    font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center')

    widths = [14, 12, 10, 18, 14, 16, 16, 14, 24]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name='modelo_vendas.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/vendas/importar-planilha', methods=['GET', 'POST'])
@login_required
def vendas_importar_planilha():
    if request.method == 'POST':
        from openpyxl import load_workbook

        arquivo = request.files.get('arquivo')
        if not arquivo:
            flash('Selecione um arquivo Excel.', 'warning')
            return redirect(url_for('vendas_importar_planilha'))

        wb = load_workbook(arquivo, data_only=True)
        ws = wb[wb.sheetnames[0]]
        linhas_importadas = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or all(v is None or str(v).strip() == '' for v in row):
                continue

            data_venda = _parse_excel_date(row[0])
            caixa_importacao = _obter_caixa_para_data_importacao(data_venda, session.get('turno'))
            tipo = str(row[1] or 'MESA').strip().upper()
            numero = int(float(row[2] or 1))
            total = parse_moeda(row[3] or 0)
            pagamento = parse_moeda(row[4] or total)
            forma_nome = row[5] or 'DINHEIRO'
            bandeira_nome = row[6] or None
            emitiu_nota = str(row[7] or '').strip().upper() in ['SIM', 'S', 'TRUE', '1', 'ON']
            observacao = row[8] or ''

            venda = Venda(
                caixa_id=caixa_importacao.id,
                tipo=tipo if tipo in ['MESA', 'BALCAO'] else 'MESA',
                numero=numero,
                total=total,
                emitiu_nota=emitiu_nota,
                observacao=observacao,
                data_hora=data_venda
            )
            db.session.add(venda)
            db.session.flush()

            forma = _lookup_or_create_forma_pagamento(forma_nome)
            bandeira = _lookup_or_create_bandeira(bandeira_nome) if bandeira_nome else None
            db.session.add(PagamentoVenda(
                venda_id=venda.id,
                forma_pagamento_id=forma.id if forma else None,
                bandeira_id=bandeira.id if bandeira else None,
                valor=pagamento,
                observacao='Importado por planilha'
            ))
            linhas_importadas += 1

        db.session.commit()
        flash(f'{linhas_importadas} venda(s) importada(s) com sucesso.', 'success')
        return redirect(url_for('vendas'))

    return render_template(
        'importar_planilha.html',
        titulo='Importar Vendas',
        subtitulo='Baixe o modelo, preencha e envie a planilha.',
        download_url=url_for('vendas_modelo_planilha'),
        upload_url=url_for('vendas_importar_planilha'),
        campos=['DATA', 'TIPO', 'N°', 'TOTAL DE MESA', 'PAGAMENTO', 'FORMA', 'BANDEIRA', 'NOTA FISCAL', 'OBSERVAÇÃO']
    )


@app.route('/delivery/modelo-planilha')
@login_required
def delivery_modelo_planilha():
    from io import BytesIO
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Delivery'

    headers = ['DATA', 'NOME', 'TOTAL', 'TAXA ENTREGA', 'PAGAMENTO', 'FORMA', 'BANDEIRA', 'NOTA FISCAL', 'MOTOBOY', 'OBSERVAÇÃO']
    ws.append(headers)
    ws.append(['07/05/2026', 'VALDECIR', 65.50, 0.00, 65.50, 'PIX', '', 'NAO', 'Moto 1', 'Exemplo'])

    fill = PatternFill('solid', fgColor='1F3B73')
    font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center')

    widths = [14, 22, 14, 16, 14, 16, 16, 14, 16, 24]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        as_attachment=True,
        download_name='modelo_delivery.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/delivery/importar-planilha', methods=['GET', 'POST'])
@login_required
def delivery_importar_planilha():
    if request.method == 'POST':
        from openpyxl import load_workbook

        arquivo = request.files.get('arquivo')
        if not arquivo:
            flash('Selecione um arquivo Excel.', 'warning')
            return redirect(url_for('delivery_importar_planilha'))

        wb = load_workbook(arquivo, data_only=True)
        ws = wb[wb.sheetnames[0]]
        linhas_importadas = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or all(v is None or str(v).strip() == '' for v in row):
                continue

            data_delivery = _parse_excel_date(row[0])
            caixa_importacao = _obter_caixa_para_data_importacao(data_delivery, session.get('turno'))
            cliente = str(row[1] or '').strip()
            total = parse_moeda(row[2] or 0)
            taxa_entrega = parse_moeda(row[3] or 0)
            pagamento = parse_moeda(row[4] or (total + taxa_entrega))
            forma_nome = row[5] or 'DINHEIRO'
            bandeira_nome = row[6] or None
            emitiu_nota = str(row[7] or '').strip().upper() in ['SIM', 'S', 'TRUE', '1', 'ON']
            motoboy_nome = row[8] or None
            observacao = row[9] or ''

            delivery = Delivery(
                caixa_id=caixa_importacao.id,
                cliente=cliente,
                total=total,
                taxa_entrega=taxa_entrega,
                motoboy_id=_lookup_or_create_motoboy(motoboy_nome).id if motoboy_nome else None,
                emitiu_nota=emitiu_nota,
                observacao=observacao,
                data_hora=data_delivery
            )
            db.session.add(delivery)
            db.session.flush()

            forma = _lookup_or_create_forma_pagamento(forma_nome)
            bandeira = _lookup_or_create_bandeira(bandeira_nome) if bandeira_nome else None
            db.session.add(PagamentoDelivery(
                delivery_id=delivery.id,
                forma_pagamento_id=forma.id if forma else None,
                bandeira_id=bandeira.id if bandeira else None,
                valor=pagamento,
                observacao='Importado por planilha'
            ))
            linhas_importadas += 1

        db.session.commit()
        flash(f'{linhas_importadas} delivery(s) importado(s) com sucesso.', 'success')
        return redirect(url_for('delivery'))

    return render_template(
        'importar_planilha.html',
        titulo='Importar Delivery',
        subtitulo='Baixe o modelo, preencha e envie a planilha.',
        download_url=url_for('delivery_modelo_planilha'),
        upload_url=url_for('delivery_importar_planilha'),
        campos=['DATA', 'NOME', 'TOTAL', 'TAXA ENTREGA', 'PAGAMENTO', 'FORMA', 'BANDEIRA', 'NOTA FISCAL', 'MOTOBOY', 'OBSERVAÇÃO']
    )
# ==================== ROUTES - DESPESAS ====================

@app.route('/despesas')
@login_required
def despesas():
    caixa = db.session.get(Caixa, session['caixa_id'])
    if not caixa:
        session.clear()
        flash('Caixa nÃ£o encontrado. Por favor, faÃ§a login novamente.', 'warning')
        return redirect(url_for('login'))

    despesas_list = Despesa.query.filter_by(caixa_id=session['caixa_id']).order_by(Despesa.data_hora.desc()).all()
    categorias = CategoriaDespesa.query.filter_by(ativo=True).all()
    formas_pagamento = FormaPagamento.query.filter_by(ativo=True).all()
    
    totais = calcular_totais_caixa(caixa)
    
    return render_template('despesas.html',
                         despesas=despesas_list,
                         categorias=categorias,
                         formas_pagamento=formas_pagamento,
                         totais=totais,
                         caixa=caixa)

@app.route('/despesas/nova', methods=['POST'])
@login_required
def nova_despesa():
    try:
        tipo = request.form.get('tipo')
        categoria_id = request.form.get('categoria_id')
        descricao = request.form.get('descricao')
        valor = parse_moeda(request.form.get('valor'))
        forma_pagamento_id = request.form.get('forma_pagamento_id')
        data_vencimento = request.form.get('data_vencimento')
        observacao = request.form.get('observacao', '')
        
        despesa = Despesa(
            caixa_id=session['caixa_id'],
            tipo=tipo,
            categoria_id=int(categoria_id) if categoria_id else None,
            descricao=descricao,
            valor=valor,
            forma_pagamento_id=int(forma_pagamento_id) if forma_pagamento_id else None,
            data_vencimento=datetime.strptime(data_vencimento, '%Y-%m-%d').date() if data_vencimento else None,
            observacao=observacao
        )
        db.session.add(despesa)
        db.session.commit()
        
        flash('Despesa registrada com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao registrar despesa: {str(e)}', 'danger')
    
    return redirect(url_for('despesas'))

# ==================== ROUTES - SANGRIA ====================

@app.route('/sangria')
@login_required
def sangria():
    caixa = db.session.get(Caixa, session['caixa_id'])
    if not caixa:
        session.clear()
        flash('Caixa nÃ£o encontrado. Por favor, faÃ§a login novamente.', 'warning')
        return redirect(url_for('login'))

    sangrias_list = Sangria.query.filter_by(caixa_id=session['caixa_id']).order_by(Sangria.data_hora.desc()).all()
    
    totais = calcular_totais_caixa(caixa)
    
    return render_template('sangria.html',
                         sangrias=sangrias_list,
                         totais=totais,
                         caixa=caixa)

@app.route('/sangria/nova', methods=['POST'])
@login_required
def nova_sangria():
    try:
        valor = parse_moeda(request.form.get('valor'))
        motivo = request.form.get('motivo')
        observacao = request.form.get('observacao', '')
        
        sangria_obj = Sangria(
            caixa_id=session['caixa_id'],
            valor=valor,
            motivo=motivo,
            observacao=observacao
        )
        db.session.add(sangria_obj)
        db.session.commit()
        
        flash('Sangria registrada com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao registrar sangria: {str(e)}', 'danger')
    
    return redirect(url_for('sangria'))

# ==================== ROUTES - ESTOQUE ====================

@app.route('/estoque')
@login_required
def estoque():
    produtos = Produto.query.filter_by(ativo=True).all()
    
    # Calcular totais
    total_produtos = len(produtos)
    total_valor = sum(p.quantidade * p.custo for p in produtos)
    criticos = len([p for p in produtos if p.quantidade <= p.estoque_minimo * 0.3])
    baixos = len([p for p in produtos if p.quantidade > p.estoque_minimo * 0.3 and p.quantidade <= p.estoque_minimo])
    
    return render_template('estoque.html',
                         produtos=produtos,
                         total_produtos=total_produtos,
                         total_valor=total_valor,
                         criticos=criticos,
                         baixos=baixos)

@app.route('/estoque/produto/novo', methods=['POST'])
@login_required
def novo_produto():
    try:
        codigo = request.form.get('codigo')
        nome = request.form.get('nome')
        categoria = request.form.get('categoria')
        custo = parse_moeda(request.form.get('custo', 0))
        preco_venda = parse_moeda(request.form.get('preco_venda', 0))
        quantidade = int(request.form.get('quantidade', 0))
        estoque_minimo = int(request.form.get('estoque_minimo', 0))
        estoque_maximo = int(request.form.get('estoque_maximo', 0))
        
        produto = Produto(
            codigo=codigo or f'PROD{Produto.query.count() + 1:03d}',
            nome=nome,
            categoria=categoria,
            custo=custo,
            preco_venda=preco_venda,
            quantidade=quantidade,
            estoque_minimo=estoque_minimo,
            estoque_maximo=estoque_maximo
        )
        db.session.add(produto)
        db.session.flush()
        
        # Criar movimentaÃ§Ã£o inicial se houver quantidade
        if quantidade > 0:
            movimentacao = MovimentacaoEstoque(
                produto_id=produto.id,
                tipo='ENTRADA',
                quantidade=quantidade,
                valor_unitario=custo,
                valor_total=custo * quantidade,
                motivo='CADASTRO',
                observacao='Cadastro inicial',
                usuario_id=session['user_id']
            )
            db.session.add(movimentacao)
        
        db.session.commit()
        flash('Produto cadastrado com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao cadastrar produto: {str(e)}', 'danger')
    
    return redirect(url_for('estoque'))

@app.route('/estoque/movimentacao/nova', methods=['POST'])
@login_required
def nova_movimentacao():
    try:
        produto_id = int(request.form.get('produto_id'))
        tipo = request.form.get('tipo')
        quantidade = int(request.form.get('quantidade'))
        valor_unitario = parse_moeda(request.form.get('valor_unitario', 0))
        motivo = request.form.get('motivo')
        observacao = request.form.get('observacao', '')
        
        produto = db.session.get(Produto, produto_id)
        if not produto:
            flash('Produto nÃ£o encontrado!', 'danger')
            return redirect(url_for('estoque'))
        
        # Verificar estoque para saÃ­das
        if tipo == 'SAIDA' and quantidade > produto.quantidade:
            flash(f'Quantidade insuficiente! DisponÃ­vel: {produto.quantidade}', 'danger')
            return redirect(url_for('estoque'))
        
        # Atualizar quantidade
        if tipo == 'ENTRADA':
            produto.quantidade += quantidade
        elif tipo == 'SAIDA':
            produto.quantidade -= quantidade
        elif tipo == 'AJUSTE':
            produto.quantidade = quantidade
        
        # Criar movimentaÃ§Ã£o
        movimentacao = MovimentacaoEstoque(
            produto_id=produto_id,
            tipo=tipo,
            quantidade=quantidade,
            valor_unitario=valor_unitario,
            valor_total=valor_unitario * quantidade,
            motivo=motivo,
            observacao=observacao,
            usuario_id=session['user_id']
        )
        db.session.add(movimentacao)
        db.session.commit()
        
        flash('MovimentaÃ§Ã£o registrada com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao registrar movimentaÃ§Ã£o: {str(e)}', 'danger')
    
    return redirect(url_for('estoque'))

# ==================== ROUTES - COMPRAS ====================

@app.route('/compras')
@login_required
def compras():
    compras = Compra.query.order_by(Compra.data.desc()).all()
    return render_template('compras.html', compras=compras)

@app.route('/compras/importar', methods=['GET', 'POST'])
@login_required
def compras_importar():
    if request.method == 'POST':
        import base64
        import io

        if request.form.get('confirm') == '1':
            xml_b64 = request.form.get('xml_b64')
            if not xml_b64:
                flash('XML nao encontrado para confirmacao.', 'danger')
                return redirect(url_for('compras_importar'))

            try:
                xml_bytes = base64.b64decode(xml_b64)
                itens = parse_nfe_xml(io.BytesIO(xml_bytes))
            except Exception as e:
                flash(f'Erro ao processar XML: {e}', 'danger')
                return redirect(url_for('compras_importar'))

            if not itens:
                flash('Nao foi possivel extrair itens do XML.', 'danger')
                return redirect(url_for('compras_importar'))

            compra = Compra(tipo='NFE', fornecedor=request.form.get('fornecedor', '-'))
            db.session.add(compra)
            total = 0
            conversao = request.form.get('conversao', '1')
            try:
                conversao = float(str(conversao).replace(',', '.'))
            except Exception:
                conversao = 1

            for it in itens:
                quantidade = float(it['quantidade'] or 0) * conversao
                preco_unit = float(it['preco_unitario'] or 0)
                item = CompraItem(
                    compra=compra,
                    nome=it['nome'][:200],
                    codigo=it.get('codigo'),
                    quantidade=quantidade,
                    unidade=it.get('unidade', 'UN'),
                    preco_unitario=preco_unit
                )
                db.session.add(item)
                total += quantidade * preco_unit
                _find_or_create_produto(it['nome'], it.get('codigo'), preco_unit, quantidade, it.get('unidade', 'UN'))

            compra.total = total
            db.session.commit()
            flash(f'Importacao concluida: {len(itens)} itens.', 'success')
            return redirect(url_for('compras'))

        arquivo = request.files.get('xml_file')
        if not arquivo:
            flash('Selecione um arquivo XML.', 'warning')
            return redirect(url_for('compras_importar'))

        xml_bytes = arquivo.read()
        itens = parse_nfe_xml(io.BytesIO(xml_bytes))
        if not itens:
            flash('Nao foi possivel extrair itens do XML.', 'danger')
            return redirect(url_for('compras_importar'))

        xml_b64 = base64.b64encode(xml_bytes).decode()
        fornecedor = request.form.get('fornecedor', '-')
        conversao = request.form.get('conversao', '1')
        return render_template('compras_import_preview.html', itens=itens, xml_b64=xml_b64, fornecedor=fornecedor, conversao=conversao)

    return render_template('compras_import.html')

@app.route('/compras/novo', methods=['GET', 'POST'])
@login_required
def compras_novo():
    if request.method == 'POST':
        fornecedor = request.form.get('fornecedor')
        compra = Compra(tipo='BALCAO', fornecedor=fornecedor)
        db.session.add(compra)

        total = 0
        nomes = request.form.getlist('nome[]')
        codigos = request.form.getlist('codigo[]')
        quantidades = request.form.getlist('quantidade[]')
        precos = request.form.getlist('preco[]')
        conversoes = request.form.getlist('conversao[]')

        for i, nome in enumerate(nomes):
            nome = (nome or '').strip()
            if not nome:
                continue
            try:
                q = float(str(quantidades[i]).replace(',', '.'))
            except Exception:
                q = 0
            try:
                p = float(str(precos[i]).replace(',', '.'))
            except Exception:
                p = 0
            try:
                conv = float(str(conversoes[i]).replace(',', '.'))
            except Exception:
                conv = 1

            quantidade = q * conv
            preco_unit = p / (conv or 1)
            item = CompraItem(
                compra=compra,
                nome=nome[:200],
                codigo=(codigos[i] or None),
                quantidade=quantidade,
                unidade='UN',
                preco_unitario=preco_unit
            )
            db.session.add(item)
            total += quantidade * preco_unit
            _find_or_create_produto(nome, codigos[i] or None, preco_unit, quantidade)

        compra.total = total
        db.session.commit()
        flash('Compra registrada com sucesso e estoque atualizado.', 'success')
        return redirect(url_for('compras'))

    return render_template('compras_novo.html')

@app.route('/compras/<int:compra_id>')
@login_required
def compra_detalhes(compra_id):
    compra = db.session.get(Compra, compra_id)
    if not compra:
        flash('Compra nao encontrada.', 'warning')
        return redirect(url_for('compras'))
    return render_template('compra_detalhes.html', compra=compra)

# ==================== ROUTES - PRODUTOS ====================

@app.route('/produtos')
@login_required
def produtos():
    produtos = Produto.query.order_by(Produto.nome.asc()).all()
    return render_template('produtos.html', produtos=produtos)

@app.route('/produtos/novo', methods=['GET', 'POST'])
@login_required
def produtos_novo():
    if request.method == 'POST':
        try:
            produto = Produto(
                codigo=request.form.get('codigo') or f'PROD{Produto.query.count() + 1:03d}',
                nome=request.form.get('nome'),
                categoria=request.form.get('categoria'),
                unidade=request.form.get('unidade') or 'UN',
                tamanho=request.form.get('tamanho') if hasattr(Produto, 'tamanho') else None,
                quantidade=int(request.form.get('quantidade', 0) or 0),
                estoque_minimo=int(request.form.get('estoque_minimo', 0) or 0),
                estoque_maximo=int(request.form.get('estoque_maximo', 0) or 0),
                custo=parse_moeda(request.form.get('custo', 0)),
                preco_venda=parse_moeda(request.form.get('preco_venda', 0)),
                ativo='ativo' in request.form
            )
            db.session.add(produto)
            db.session.commit()
            flash('Produto cadastrado com sucesso!', 'success')
            return redirect(url_for('produtos'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao cadastrar produto: {e}', 'danger')
    return render_template('produto_form.html', produto=None)

@app.route('/produtos/<int:produto_id>/editar', methods=['GET', 'POST'])
@login_required
def produtos_editar(produto_id):
    produto = db.session.get(Produto, produto_id)
    if not produto:
        flash('Produto nao encontrado.', 'warning')
        return redirect(url_for('produtos'))

    if request.method == 'POST':
        try:
            produto.codigo = request.form.get('codigo')
            produto.nome = request.form.get('nome')
            produto.categoria = request.form.get('categoria')
            produto.unidade = request.form.get('unidade') or 'UN'
            produto.quantidade = int(request.form.get('quantidade', produto.quantidade or 0) or 0)
            produto.estoque_minimo = int(request.form.get('estoque_minimo', produto.estoque_minimo or 0) or 0)
            produto.estoque_maximo = int(request.form.get('estoque_maximo', produto.estoque_maximo or 0) or 0)
            produto.custo = parse_moeda(request.form.get('custo', produto.custo))
            produto.preco_venda = parse_moeda(request.form.get('preco_venda', produto.preco_venda))
            produto.ativo = 'ativo' in request.form
            db.session.commit()
            flash('Produto atualizado com sucesso!', 'success')
            return redirect(url_for('produtos'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar produto: {e}', 'danger')

    return render_template('produto_form.html', produto=produto)

# ==================== ROUTES - DASHBOARD ====================

@app.route('/dashboard')
@login_required
@dashboard_required 
def dashboard():
    # Pegar filtros
    periodo = request.args.get('periodo', 'month')
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    turno = request.args.get('turno', 'all')
    
    # Definir datas
    hoje = datetime.now().date()
    if periodo == 'today':
        data_inicio = hoje
        data_fim = hoje
    elif periodo == 'week':
        data_inicio = hoje - timedelta(days=7)
        data_fim = hoje
    elif periodo == 'month':
        data_inicio = hoje.replace(day=1)
        data_fim = hoje
    else:
        if data_inicio:
            data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        if data_fim:
            data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
    
    # Buscar caixas no perÃ­odo
    caixas_query = Caixa.query.filter(
        Caixa.data >= data_inicio,
        Caixa.data <= data_fim
    )

    if turno and turno != 'all':
        caixas_query = caixas_query.filter(Caixa.turno == turno)

    caixas = caixas_query.all()
    
    # Calcular mÃ©tricas
    metricas = calcular_metricas_dashboard(caixas)
    
    # Calcular mÃ©tricas avanÃ§adas
    metricas_avancadas = calcular_metricas_avancadas(caixas)
    
    return render_template('dashboard.html',
                         metricas=metricas,
                         metricas_avancadas=metricas_avancadas,
                         caixas=caixas,
                         periodo=periodo,
                         data_inicio=data_inicio,
                         data_fim=data_fim,
                         turno=turno)

# ==================== ROUTES - CONFIGURAÃ‡Ã•ES ====================

@app.route('/configuracoes')
@admin_required
def configuracoes():
    usuarios = Usuario.query.all()
    formas_pagamento = FormaPagamento.query.all()
    bandeiras = BandeiraCartao.query.all()
    categorias = CategoriaDespesa.query.all()
    motoboys = Motoboy.query.all()
    
    return render_template('configuracoes.html',
                         usuarios=usuarios,
                         formas_pagamento=formas_pagamento,
                         bandeiras=bandeiras,
                         categorias=categorias,
                         motoboys=motoboys)

@app.route('/configuracoes/usuario/novo', methods=['POST'])
@admin_required
def novo_usuario():
    try:
        nome = request.form.get('nome')
        senha = request.form.get('senha')
        perfil = request.form.get('perfil', 'OPERADOR')
        acesso_dashboard = 'acesso_dashboard' in request.form
        acesso_configuracoes = 'acesso_configuracoes' in request.form
        acesso_relatorios = 'acesso_relatorios' in request.form
        
        # Verificar se usuÃ¡rio jÃ¡ existe
        usuario_existe = Usuario.query.filter_by(nome=nome).first()
        if usuario_existe:
            flash('JÃ¡ existe um usuÃ¡rio com este nome!', 'warning')
            return redirect(url_for('configuracoes'))
        
        usuario = Usuario(
            nome=nome,
            senha=generate_password_hash(senha),
            perfil=perfil,
            acesso_dashboard=acesso_dashboard,
            acesso_configuracoes=acesso_configuracoes,
            acesso_relatorios=acesso_relatorios
        )
        db.session.add(usuario)
        db.session.commit()
        
        flash(f'UsuÃ¡rio {nome} criado com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao criar usuÃ¡rio: {str(e)}', 'danger')
    
    return redirect(url_for('configuracoes'))

@app.route('/configuracoes/forma-pagamento/nova', methods=['POST'])
@admin_required
def nova_forma_pagamento():
    try:
        nome = request.form.get('nome')
        
        if not nome or nome.strip() == '':
            flash('Informe o nome da forma de pagamento!', 'warning')
            return redirect(url_for('configuracoes'))
        
        # Verificar se jÃ¡ existe
        existe = FormaPagamento.query.filter_by(nome=nome).first()
        if existe:
            flash('Esta forma de pagamento jÃ¡ existe!', 'warning')
            return redirect(url_for('configuracoes'))
        
        forma = FormaPagamento(nome=nome)
        db.session.add(forma)
        db.session.commit()
        
        flash(f'Forma de pagamento "{nome}" criada com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao criar forma de pagamento: {str(e)}', 'danger')
    
    return redirect(url_for('configuracoes'))

@app.route('/configuracoes/bandeira/nova', methods=['POST'])
@admin_required
def nova_bandeira():
    try:
        nome = request.form.get('nome')
        
        if not nome or nome.strip() == '':
            flash('Informe o nome da bandeira!', 'warning')
            return redirect(url_for('configuracoes'))
        
        # Verificar se jÃ¡ existe
        existe = BandeiraCartao.query.filter_by(nome=nome).first()
        if existe:
            flash('Esta bandeira jÃ¡ existe!', 'warning')
            return redirect(url_for('configuracoes'))
        
        bandeira = BandeiraCartao(nome=nome)
        db.session.add(bandeira)
        db.session.commit()
        
        flash(f'Bandeira "{nome}" criada com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao criar bandeira: {str(e)}', 'danger')
    
    return redirect(url_for('configuracoes'))

@app.route('/configuracoes/categoria/nova', methods=['POST'])
@admin_required
def nova_categoria():
    try:
        nome = request.form.get('nome')
        tipo = request.form.get('tipo')
        
        if not nome or nome.strip() == '':
            flash('Informe o nome da categoria!', 'warning')
            return redirect(url_for('configuracoes'))
        
        # Verificar se jÃ¡ existe
        existe = CategoriaDespesa.query.filter_by(nome=nome, tipo=tipo).first()
        if existe:
            flash('Esta categoria jÃ¡ existe!', 'warning')
            return redirect(url_for('configuracoes'))
        
        categoria = CategoriaDespesa(nome=nome, tipo=tipo)
        db.session.add(categoria)
        db.session.commit()
        
        flash(f'Categoria "{nome}" ({tipo}) criada com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao criar categoria: {str(e)}', 'danger')
    
    return redirect(url_for('configuracoes'))

@app.route('/configuracoes/motoboy/novo', methods=['POST'])
@admin_required
def novo_motoboy():
    try:
        nome = request.form.get('nome')
        taxa_padrao = parse_moeda(request.form.get('taxa_padrao', 5.00))
        
        if not nome or nome.strip() == '':
            flash('Informe o nome do motoboy!', 'warning')
            return redirect(url_for('configuracoes'))
        
        # Verificar se jÃ¡ existe
        existe = Motoboy.query.filter_by(nome=nome).first()
        if existe:
            flash('JÃ¡ existe um motoboy com este nome!', 'warning')
            return redirect(url_for('configuracoes'))
        
        motoboy = Motoboy(nome=nome, taxa_padrao=taxa_padrao)
        db.session.add(motoboy)
        db.session.commit()
        
        flash(f'Motoboy "{nome}" cadastrado com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao cadastrar motoboy: {str(e)}', 'danger')
    
    return redirect(url_for('configuracoes'))

# ==================== ROUTES - GESTÃƒO DE CAIXAS (ADMIN) ====================

@app.route('/admin/caixas')
@admin_required
def admin_caixas():
    """Lista todos os caixas para administraÃ§Ã£o"""
    page = request.args.get('page', 1, type=int)
    status_filter = request.args.get('status', 'all')
    
    query = Caixa.query.order_by(Caixa.data.desc(), Caixa.hora_abertura.desc())
    
    if status_filter != 'all':
        query = query.filter_by(status=status_filter.upper())
    
    caixas = query.paginate(page=page, per_page=20, error_out=False)
    
    return render_template('admin_caixas.html', caixas=caixas, status_filter=status_filter)

@app.route('/admin/caixa/<int:caixa_id>/visualizar')
@admin_required
def admin_visualizar_caixa(caixa_id):
    """Visualizar detalhes de um caixa especÃ­fico"""
    caixa = db.session.get(Caixa, caixa_id)
    if not caixa:
        flash('Caixa nÃ£o encontrado!', 'danger')
        return redirect(url_for('admin_caixas'))
    
    totais = calcular_totais_fechamento(caixa)
    
    return render_template('admin_visualizar_caixa.html', caixa=caixa, totais=totais)

@app.route('/admin/caixa/<int:caixa_id>/excluir-completo', methods=['POST'])
@admin_required
def admin_excluir_caixa_completo(caixa_id):
    """Excluir um caixa e TODOS os seus registros (vendas, deliveries, despesas, etc.)"""
    try:
        caixa = db.session.get(Caixa, caixa_id)
        if not caixa:
            flash('Caixa nÃ£o encontrado!', 'danger')
            return redirect(url_for('admin_caixas'))
        
        # Registrar informaÃ§Ãµes para mensagem
        info_caixa = f"Caixa #{caixa.id} - {caixa.data.strftime('%d/%m/%Y')} - {caixa.turno}"
        
        # Excluir todos os registros relacionados (em ordem para evitar constraints)
        
        # 1. Excluir pagamentos das vendas
        for venda in caixa.vendas:
            PagamentoVenda.query.filter_by(venda_id=venda.id).delete()
        
        # 2. Excluir pagamentos dos deliveries
        for delivery in caixa.deliveries:
            PagamentoDelivery.query.filter_by(delivery_id=delivery.id).delete()
        
        # 3. Excluir vendas
        Venda.query.filter_by(caixa_id=caixa_id).delete()
        
        # 4. Excluir deliveries
        Delivery.query.filter_by(caixa_id=caixa_id).delete()
        
        # 5. Excluir despesas
        Despesa.query.filter_by(caixa_id=caixa_id).delete()
        
        # 6. Excluir sangrias
        Sangria.query.filter_by(caixa_id=caixa_id).delete()
        
        # 7. Excluir suprimentos (se existir)
        if hasattr(caixa, 'suprimentos'):
            Suprimento.query.filter_by(caixa_id=caixa_id).delete()
        
        # 8. Finalmente, excluir o caixa
        db.session.delete(caixa)
        db.session.commit()
        
        flash(f'OK: Caixa {info_caixa} excluído permanentemente com todos os registros!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir caixa: {str(e)}', 'danger')
    
    return redirect(url_for('admin_caixas'))

@app.route('/admin/caixa/<int:caixa_id>/gerar-pdf')
@admin_required
def admin_gerar_pdf_caixa(caixa_id):
    """Gerar PDF do fechamento do caixa"""
    try:
        from flask import make_response
        
        caixa = db.session.get(Caixa, caixa_id)
        if not caixa:
            flash('Caixa nÃ£o encontrado!', 'danger')
            return redirect(url_for('admin_caixas'))
        
        totais = calcular_totais_fechamento(caixa)
        
        # Renderizar HTML do relatÃ³rio
        html = render_template('relatorio_imprimivel.html', 
                              caixa=caixa, 
                              totais=totais,
                              now=datetime.now)
        
        # Tentar importar weasyprint para gerar PDF
        try:
            from weasyprint import HTML
            pdf = HTML(string=html).write_pdf()
            
            response = make_response(pdf)
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = f'inline; filename=caixa_{caixa_id}_{caixa.data.strftime("%d%m%Y")}_{caixa.turno}.pdf'
            
            return response
        except Exception as e:
            # Se houver erro (falta de DLLs GTK ou biblioteca nÃ£o instalada), retorna HTML
            flash(f'Modo de compatibilidade ativado: Imprima usando o navegador (Ctrl+P).', 'warning')
            return html
            
    except Exception as e:
        flash(f'Erro ao gerar PDF: {str(e)}', 'danger')
        return redirect(url_for('admin_visualizar_caixa', caixa_id=caixa_id))

@app.route('/admin/caixa/<int:caixa_id>/editar', methods=['GET', 'POST'])
@admin_required
def admin_editar_caixa(caixa_id):
    """Editar um caixa (apenas admin)"""
    caixa = db.session.get(Caixa, caixa_id)
    if not caixa:
        flash('Caixa nÃ£o encontrado!', 'danger')
        return redirect(url_for('admin_caixas'))
    
    if request.method == 'POST':
        try:
            caixa.saldo_inicial = parse_moeda(request.form.get('saldo_inicial', caixa.saldo_inicial))
            caixa.saldo_final = parse_moeda(request.form.get('saldo_final', caixa.saldo_final))
            
            db.session.commit()
            flash('Caixa atualizado com sucesso!', 'success')
            return redirect(url_for('admin_visualizar_caixa', caixa_id=caixa_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar caixa: {str(e)}', 'danger')
    
    totais = calcular_totais_fechamento(caixa)
    return render_template('admin_editar_caixa.html', caixa=caixa, totais=totais)

@app.route('/admin/venda/<int:venda_id>/editar', methods=['POST'])
@admin_required
def admin_editar_venda(venda_id):
    """Editar uma venda"""
    try:
        venda = db.session.get(Venda, venda_id)
        if not venda:
            flash('Venda nÃ£o encontrada!', 'danger')
            return redirect(url_for('admin_caixas'))
        
        venda.total = parse_moeda(request.form.get('total', venda.total))
        venda.observacao = request.form.get('observacao', venda.observacao)
        
        db.session.commit()
        flash('Venda atualizada com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao atualizar venda: {str(e)}', 'danger')
    
    return redirect(url_for('admin_visualizar_caixa', caixa_id=venda.caixa_id))

@app.route('/admin/venda/<int:venda_id>/deletar', methods=['POST'])
@admin_required
def admin_deletar_venda(venda_id):
    """Deletar uma venda"""
    try:
        venda = db.session.get(Venda, venda_id)
        if not venda:
            flash('Venda nÃ£o encontrada!', 'danger')
            return redirect(url_for('admin_caixas'))
        
        caixa_id = venda.caixa_id
        
        # Deletar pagamentos relacionados
        PagamentoVenda.query.filter_by(venda_id=venda_id).delete()
        
        # Deletar venda
        db.session.delete(venda)
        db.session.commit()
        
        flash('Venda removida com sucesso!', 'success')
        return redirect(url_for('admin_visualizar_caixa', caixa_id=caixa_id))
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao remover venda: {str(e)}', 'danger')
        return redirect(url_for('admin_caixas'))

@app.route('/admin/despesa/<int:despesa_id>/deletar', methods=['POST'])
@admin_required
def admin_deletar_despesa(despesa_id):
    """Deletar uma despesa"""
    try:
        despesa = db.session.get(Despesa, despesa_id)
        if not despesa:
            flash('Despesa nÃ£o encontrada!', 'danger')
            return redirect(url_for('admin_caixas'))
        
        caixa_id = despesa.caixa_id
        
        db.session.delete(despesa)
        db.session.commit()
        
        flash('Despesa removida com sucesso!', 'success')
        return redirect(url_for('admin_visualizar_caixa', caixa_id=caixa_id))
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao remover despesa: {str(e)}', 'danger')
        return redirect(url_for('admin_caixas'))

@app.route('/admin/caixa/<int:caixa_id>/reabrir', methods=['POST'])
@admin_required
def admin_reabrir_caixa(caixa_id):
    """Reabrir um caixa fechado (apenas admin)"""
    try:
        caixa = db.session.get(Caixa, caixa_id)
        if not caixa:
            flash('Caixa nÃ£o encontrado!', 'danger')
            return redirect(url_for('admin_caixas'))
        
        if caixa.status == 'ABERTO':
            flash('Este caixa jÃ¡ estÃ¡ aberto!', 'warning')
            return redirect(url_for('admin_visualizar_caixa', caixa_id=caixa_id))
        
        # Reabrir caixa
        caixa.status = 'ABERTO'
        caixa.hora_fechamento = None
        
        db.session.commit()
        
        flash(f'Caixa #{caixa_id} reaberto com sucesso! Agora pode ser editado.', 'success')
        return redirect(url_for('admin_visualizar_caixa', caixa_id=caixa_id))
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao reabrir caixa: {str(e)}', 'danger')
        return redirect(url_for('admin_caixas'))

@app.route('/admin/caixa/<int:caixa_id>/fechar-forcado', methods=['POST'])
@admin_required
def admin_fechar_caixa_forcado(caixa_id):
    """Fechar um caixa aberto forÃ§adamente (admin)"""
    try:
        caixa = db.session.get(Caixa, caixa_id)
        if not caixa:
            flash('Caixa nÃ£o encontrado!', 'danger')
            return redirect(url_for('admin_caixas'))
        
        if caixa.status == 'FECHADO':
            flash('Este caixa jÃ¡ estÃ¡ fechado!', 'warning')
            return redirect(url_for('admin_visualizar_caixa', caixa_id=caixa_id))
        
        # Calcular saldo final
        totais = calcular_totais_fechamento(caixa)
        caixa.saldo_final = totais['saldo_final']
        caixa.status = 'FECHADO'
        caixa.hora_fechamento = datetime.utcnow()
        
        db.session.commit()
        
        flash(f'Aviso: Caixa #{caixa_id} fechado pelo administrador! Operador: {caixa.operador.nome}', 'warning')
        return redirect(url_for('admin_visualizar_caixa', caixa_id=caixa_id))
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao fechar caixa: {str(e)}', 'danger')
        return redirect(url_for('admin_caixas'))

@app.route('/admin/caixa/<int:caixa_id>/gerar-relatorio')
@admin_required
def admin_gerar_relatorio(caixa_id):
    """Gerar relatÃ³rio do caixa em formato imprimÃ­vel"""
    try:
        caixa = db.session.get(Caixa, caixa_id)
        if not caixa:
            flash('Caixa nÃ£o encontrado!', 'danger')
            return redirect(url_for('admin_caixas'))
        
        totais = calcular_totais_fechamento(caixa)
        
        # Renderizar HTML otimizado para impressÃ£o
        return render_template('relatorio_imprimivel.html', 
                             caixa=caixa, 
                             totais=totais,
                             now=datetime.now)
        
    except Exception as e:
        flash(f'Erro ao gerar relatÃ³rio: {str(e)}', 'danger')
        return redirect(url_for('admin_visualizar_caixa', caixa_id=caixa_id))

# ==================== ROTA EXPORTAR EXCEL REAL (.xlsx) ====================

@app.route('/exportar/excel-real/<int:caixa_id>')
@login_required
def exportar_excel_real(caixa_id):
    """Exportar todos os movimentos do caixa para Excel REAL (.xlsx)"""
    try:
        from io import BytesIO
        from flask import send_file
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        caixa = db.session.get(Caixa, caixa_id)
        if not caixa:
            flash('Caixa nÃ£o encontrado!', 'danger')
            return redirect(url_for('dashboard'))
        
        # Verificar permissÃ£o
        usuario = db.session.get(Usuario, session['user_id'])
        if not usuario.acesso_configuracoes and caixa.operador_id != usuario.id:
            flash('VocÃª nÃ£o tem permissÃ£o para exportar este caixa!', 'danger')
            return redirect(url_for('dashboard'))
        
        # Criar um workbook Excel
        wb = Workbook()
        
        # ========== ESTILOS ==========
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_font = Font(name='Arial', size=10)
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        number_alignment = Alignment(horizontal='right', vertical='center')
        currency_format = '"R$"#,##0.00'
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ========== ABA 1: INFORMAÃ‡Ã•ES GERAIS ==========
        ws_info = wb.active
        ws_info.title = "INFORMAÃ‡Ã•ES"
        
        # TÃ­tulo
        ws_info.merge_cells('A1:D1')
        ws_info['A1'] = f"RELATÃ“RIO DO CAIXA #{caixa.id}"
        ws_info['A1'].font = Font(name='Arial', size=14, bold=True)
        ws_info['A1'].alignment = Alignment(horizontal='center')
        
        # InformaÃ§Ãµes bÃ¡sicas
        info_data = [
            ['DATA DO CAIXA:', caixa.data.strftime('%d/%m/%Y')],
            ['TURNO:', caixa.turno],
            ['OPERADOR:', caixa.operador.nome],
            ['STATUS:', caixa.status],
            ['HORA ABERTURA:', caixa.hora_abertura.strftime('%H:%M:%S')],
            ['HORA FECHAMENTO:', caixa.hora_fechamento.strftime('%H:%M:%S') if caixa.hora_fechamento else '-'],
            ['SALDO INICIAL:', caixa.saldo_inicial],
            ['SALDO FINAL:', caixa.saldo_final if caixa.saldo_final else '-']
        ]
        
        for i, (label, value) in enumerate(info_data, start=3):
            ws_info[f'A{i}'] = label
            ws_info[f'B{i}'] = value
            ws_info[f'A{i}'].font = Font(name='Arial', size=11, bold=True)
            if isinstance(value, (int, float)):
                ws_info[f'B{i}'].number_format = currency_format
        
        # Ajustar largura das colunas
        ws_info.column_dimensions['A'].width = 25
        ws_info.column_dimensions['B'].width = 30
        
        # ========== ABA 2: VENDAS ==========
        ws_vendas = wb.create_sheet("VENDAS")
        
        # CabeÃ§alho
        headers_vendas = [
            'ID', 'DATA', 'HORA', 'TIPO', 'NÃšMERO', 'VALOR TOTAL',
            'FORMA PAGAMENTO', 'BANDEIRA', 'VALOR PAGO', 'NOTA FISCAL', 'OBSERVAÃ‡ÃƒO'
        ]
        
        for col, header in enumerate(headers_vendas, start=1):
            cell = ws_vendas.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Dados
        row = 2
        for venda in caixa.vendas:
            for pag in venda.pagamentos:
                ws_vendas.cell(row=row, column=1, value=venda.id).alignment = cell_alignment
                ws_vendas.cell(row=row, column=2, value=venda.data_hora.strftime('%d/%m/%Y')).alignment = cell_alignment
                ws_vendas.cell(row=row, column=3, value=venda.data_hora.strftime('%H:%M:%S')).alignment = cell_alignment
                ws_vendas.cell(row=row, column=4, value=venda.tipo).alignment = cell_alignment
                ws_vendas.cell(row=row, column=5, value=venda.numero if venda.numero else '-').alignment = cell_alignment
                ws_vendas.cell(row=row, column=6, value=venda.total).number_format = currency_format
                ws_vendas.cell(row=row, column=6).alignment = number_alignment
                ws_vendas.cell(row=row, column=7, value=pag.forma_pagamento.nome if pag.forma_pagamento else '-').alignment = cell_alignment
                ws_vendas.cell(row=row, column=8, value=pag.bandeira.nome if pag.bandeira else '-').alignment = cell_alignment
                ws_vendas.cell(row=row, column=9, value=pag.valor).number_format = currency_format
                ws_vendas.cell(row=row, column=9).alignment = number_alignment
                ws_vendas.cell(row=row, column=10, value='SIM' if venda.emitiu_nota else 'NÃƒO').alignment = Alignment(horizontal='center')
                ws_vendas.cell(row=row, column=11, value=venda.observacao if venda.observacao else '-').alignment = cell_alignment
                row += 1
        
        # Ajustar largura das colunas
        column_widths = [8, 12, 10, 10, 8, 12, 15, 12, 12, 12, 30]
        for i, width in enumerate(column_widths, start=1):
            ws_vendas.column_dimensions[get_column_letter(i)].width = width
        
        # Congelar cabeÃ§alho
        ws_vendas.freeze_panes = 'A2'
        
        # Ajustar altura das linhas para melhor visualizaÃ§Ã£o
        for row in range(1, ws_vendas.max_row + 1):
            ws_vendas.row_dimensions[row].height = 20
            
        # ========== ABA 3: DELIVERIES ==========
        ws_delivery = wb.create_sheet("DELIVERIES")
        
        headers_delivery = [
            'ID', 'DATA', 'HORA', 'CLIENTE', 'VALOR PEDIDO', 'TAXA ENTREGA',
            'TOTAL', 'MOTOBOY', 'FORMA PAGAMENTO', 'VALOR PAGO', 'NOTA FISCAL', 'OBSERVAÃ‡ÃƒO'
        ]
        
        for col, header in enumerate(headers_delivery, start=1):
            cell = ws_delivery.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
            cell.alignment = header_alignment
            cell.border = thin_border
        
        row = 2
        for delivery in caixa.deliveries:
            for pag in delivery.pagamentos:
                ws_delivery.cell(row=row, column=1, value=delivery.id).alignment = cell_alignment
                ws_delivery.cell(row=row, column=2, value=delivery.data_hora.strftime('%d/%m/%Y')).alignment = cell_alignment
                ws_delivery.cell(row=row, column=3, value=delivery.data_hora.strftime('%H:%M:%S')).alignment = cell_alignment
                ws_delivery.cell(row=row, column=4, value=delivery.cliente).alignment = cell_alignment
                ws_delivery.cell(row=row, column=5, value=delivery.total).number_format = currency_format
                ws_delivery.cell(row=row, column=5).alignment = number_alignment
                ws_delivery.cell(row=row, column=6, value=delivery.taxa_entrega).number_format = currency_format
                ws_delivery.cell(row=row, column=6).alignment = number_alignment
                total_com_taxa = delivery.total + delivery.taxa_entrega
                ws_delivery.cell(row=row, column=7, value=total_com_taxa).number_format = currency_format
                ws_delivery.cell(row=row, column=7).alignment = number_alignment
                ws_delivery.cell(row=row, column=8, value=delivery.motoboy.nome if delivery.motoboy else '-').alignment = cell_alignment
                ws_delivery.cell(row=row, column=9, value=pag.forma_pagamento.nome if pag.forma_pagamento else '-').alignment = cell_alignment
                ws_delivery.cell(row=row, column=10, value=pag.valor).number_format = currency_format
                ws_delivery.cell(row=row, column=10).alignment = number_alignment
                ws_delivery.cell(row=row, column=11, value='SIM' if delivery.emitiu_nota else 'NÃƒO').alignment = Alignment(horizontal='center')
                ws_delivery.cell(row=row, column=12, value=delivery.observacao if delivery.observacao else '-').alignment = cell_alignment
                row += 1
        
        column_widths_delivery = [8, 12, 10, 25, 12, 12, 12, 15, 15, 12, 12, 30]
        for i, width in enumerate(column_widths_delivery, start=1):
            ws_delivery.column_dimensions[get_column_letter(i)].width = width
        
        ws_delivery.freeze_panes = 'A2'
        
        for row in range(1, ws_delivery.max_row + 1):
            ws_delivery.row_dimensions[row].height = 20
        
        # ========== ABA 4: DESPESAS ==========
        ws_despesas = wb.create_sheet("DESPESAS")
        
        headers_despesas = [
            'ID', 'DATA', 'HORA', 'TIPO', 'CATEGORIA', 'DESCRIÃ‡ÃƒO',
            'VALOR', 'FORMA PAGAMENTO', 'OBSERVAÃ‡ÃƒO'
        ]
        
        for col, header in enumerate(headers_despesas, start=1):
            cell = ws_despesas.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
            cell.alignment = header_alignment
            cell.border = thin_border
        
        row = 2
        for despesa in caixa.despesas:
            ws_despesas.cell(row=row, column=1, value=despesa.id).alignment = cell_alignment
            ws_despesas.cell(row=row, column=2, value=despesa.data_hora.strftime('%d/%m/%Y')).alignment = cell_alignment
            ws_despesas.cell(row=row, column=3, value=despesa.data_hora.strftime('%H:%M:%S')).alignment = cell_alignment
            ws_despesas.cell(row=row, column=4, value=despesa.tipo).alignment = cell_alignment
            ws_despesas.cell(row=row, column=5, value=despesa.categoria.nome if despesa.categoria else '-').alignment = cell_alignment
            ws_despesas.cell(row=row, column=6, value=despesa.descricao).alignment = cell_alignment
            ws_despesas.cell(row=row, column=7, value=despesa.valor).number_format = currency_format
            ws_despesas.cell(row=row, column=7).alignment = number_alignment
            ws_despesas.cell(row=row, column=8, value=despesa.forma_pagamento.nome if despesa.forma_pagamento else '-').alignment = cell_alignment
            ws_despesas.cell(row=row, column=9, value=despesa.observacao if despesa.observacao else '-').alignment = cell_alignment
            row += 1
        
        column_widths_despesas = [8, 12, 10, 12, 20, 40, 12, 20, 30]
        for i, width in enumerate(column_widths_despesas, start=1):
            ws_despesas.column_dimensions[get_column_letter(i)].width = width
        
        ws_despesas.freeze_panes = 'A2'
        
        for row in range(1, ws_despesas.max_row + 1):
            ws_despesas.row_dimensions[row].height = 20
        
        # ========== ABA 5: SANGRIA ==========
        ws_sangria = wb.create_sheet("SANGRIA")
        
        headers_sangria = [
            'ID', 'DATA', 'HORA', 'MOTIVO', 'VALOR', 'OBSERVAÃ‡ÃƒO'
        ]
        
        for col, header in enumerate(headers_sangria, start=1):
            cell = ws_sangria.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            cell.alignment = header_alignment
            cell.border = thin_border
        
        row = 2
        for sangria in caixa.sangrias:
            ws_sangria.cell(row=row, column=1, value=sangria.id).alignment = cell_alignment
            ws_sangria.cell(row=row, column=2, value=sangria.data_hora.strftime('%d/%m/%Y')).alignment = cell_alignment
            ws_sangria.cell(row=row, column=3, value=sangria.data_hora.strftime('%H:%M:%S')).alignment = cell_alignment
            ws_sangria.cell(row=row, column=4, value=sangria.motivo).alignment = cell_alignment
            ws_sangria.cell(row=row, column=5, value=sangria.valor).number_format = currency_format
            ws_sangria.cell(row=row, column=5).alignment = number_alignment
            ws_sangria.cell(row=row, column=6, value=sangria.observacao if sangria.observacao else '-').alignment = cell_alignment
            row += 1
        
        column_widths_sangria = [8, 12, 10, 30, 12, 40]
        for i, width in enumerate(column_widths_sangria, start=1):
            ws_sangria.column_dimensions[get_column_letter(i)].width = width
        
        ws_sangria.freeze_panes = 'A2'
        
        for row in range(1, ws_sangria.max_row + 1):
            ws_sangria.row_dimensions[row].height = 20
        
        # ========== ABA 6: RESUMO FINANCEIRO ==========
        ws_resumo = wb.create_sheet("RESUMO")
        
        totais = calcular_totais_fechamento(caixa)
        
        resumo_data = [
            ['DESCRIÃ‡ÃƒO', 'VALOR (R$)'],
            ['SALDO INICIAL', caixa.saldo_inicial],
            ['VENDAS MESA/BALCÃƒO', totais['vendas_loja']],
            ['VENDAS DELIVERY', totais['vendas_delivery']],
            ['TOTAL VENDAS', totais['total_vendas']],
            ['TOTAL DESPESAS', totais['despesas']],
            ['TOTAL SANGRIA', totais['sangrias']],
            ['SALDO FINAL', totais['saldo_final']]
        ]
        
        for r_idx, row_data in enumerate(resumo_data, start=1):
            for c_idx, cell_data in enumerate(row_data, start=1):
                cell = ws_resumo.cell(row=r_idx, column=c_idx, value=cell_data)
                
                if r_idx == 1:  # CabeÃ§alho
                    cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
                    cell.fill = PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.font = Font(name='Arial', size=10)
                    if c_idx == 2:  # Coluna de valores
                        cell.number_format = currency_format
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                
                cell.border = thin_border
        
        # Destacar linha do saldo final
        ws_resumo.cell(row=8, column=1).font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        ws_resumo.cell(row=8, column=1).fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
        ws_resumo.cell(row=8, column=2).font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        ws_resumo.cell(row=8, column=2).fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
        
        ws_resumo.column_dimensions['A'].width = 30
        ws_resumo.column_dimensions['B'].width = 20
        
        # ========== SALVAR EM MEMÃ“RIA ==========
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Gerar nome do arquivo
        filename = f"caixa_{caixa_id}_{caixa.data.strftime('%d%m%Y')}_{caixa.turno}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f'Erro ao exportar Excel: {str(e)}', 'danger')
        import traceback
        print(traceback.format_exc())
        return redirect(url_for('admin_visualizar_caixa', caixa_id=caixa_id))

# ==================== ROUTES - EDIÃ‡ÃƒO DETALHADA (ADMIN) ====================

@app.route('/admin/venda/<int:venda_id>/editar-detalhes', methods=['GET', 'POST'])
@admin_required
def admin_editar_venda_detalhes(venda_id):
    """Editar venda completa com pagamentos"""
    venda = db.session.get(Venda, venda_id)
    if not venda:
        flash('Venda nÃ£o encontrada!', 'danger')
        return redirect(url_for('admin_caixas'))
    
    if request.method == 'POST':
        try:
            venda.tipo = request.form.get('tipo', venda.tipo)
            venda.numero = request.form.get('numero', venda.numero)
            venda.total = parse_moeda(request.form.get('total', venda.total))
            venda.observacao = request.form.get('observacao', venda.observacao)
            venda.emitiu_nota = 'emitiu_nota' in request.form
            
            # Atualizar pagamentos
            PagamentoVenda.query.filter_by(venda_id=venda_id).delete()
            
            forma_ids = request.form.getlist('forma_pagamento_id[]')
            valores = request.form.getlist('valor[]')
            bandeira_ids = request.form.getlist('bandeira_id[]')
            
            for i in range(len(forma_ids)):
                if forma_ids[i] and valores[i]:
                    pagamento = PagamentoVenda(
                        venda_id=venda_id,
                        forma_pagamento_id=int(forma_ids[i]),
                        valor=parse_moeda(valores[i]),
                        bandeira_id=int(bandeira_ids[i]) if bandeira_ids[i] and bandeira_ids[i] != '' else None
                    )
                    db.session.add(pagamento)
            
            db.session.commit()
            flash('Venda atualizada com sucesso!', 'success')
            return redirect(url_for('admin_visualizar_caixa', caixa_id=venda.caixa_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar venda: {str(e)}', 'danger')
    
    formas_pagamento = FormaPagamento.query.all()
    bandeiras = BandeiraCartao.query.all()
    
    return render_template('admin_editar_venda_detalhes.html', 
                         venda=venda, 
                         formas_pagamento=formas_pagamento,
                         bandeiras=bandeiras)

@app.route('/admin/delivery/<int:delivery_id>/editar-detalhes', methods=['GET', 'POST'])
@admin_required
def admin_editar_delivery_detalhes(delivery_id):
    """Editar delivery completo com pagamentos"""
    delivery = db.session.get(Delivery, delivery_id)
    if not delivery:
        flash('Delivery nÃ£o encontrado!', 'danger')
        return redirect(url_for('admin_caixas'))
    
    if request.method == 'POST':
        try:
            delivery.cliente = request.form.get('cliente', delivery.cliente)
            if hasattr(delivery, 'endereco'):
                delivery.endereco = request.form.get('endereco', delivery.endereco)
            if hasattr(delivery, 'telefone'):
                delivery.telefone = request.form.get('telefone', delivery.telefone)
            delivery.total = parse_moeda(request.form.get('total', delivery.total))
            delivery.taxa_entrega = parse_moeda(request.form.get('taxa_entrega', delivery.taxa_entrega))
            delivery.observacao = request.form.get('observacao', delivery.observacao)
            delivery.emitiu_nota = 'emitiu_nota' in request.form
            delivery.motoboy_id = int(request.form.get('motoboy_id')) if request.form.get('motoboy_id') else None
            
            # Atualizar pagamentos
            PagamentoDelivery.query.filter_by(delivery_id=delivery_id).delete()
            
            forma_ids = request.form.getlist('forma_pagamento_id[]')
            valores = request.form.getlist('valor[]')
            
            for i in range(len(forma_ids)):
                if forma_ids[i] and valores[i]:
                    pagamento = PagamentoDelivery(
                        delivery_id=delivery_id,
                        forma_pagamento_id=int(forma_ids[i]),
                        valor=parse_moeda(valores[i])
                    )
                    db.session.add(pagamento)
            
            db.session.commit()
            flash('Delivery atualizado com sucesso!', 'success')
            return redirect(url_for('admin_visualizar_caixa', caixa_id=delivery.caixa_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar delivery: {str(e)}', 'danger')
    
    formas_pagamento = FormaPagamento.query.all()
    motoboys = Motoboy.query.all()
    
    return render_template('admin_editar_delivery_detalhes.html', 
                         delivery=delivery,
                         formas_pagamento=formas_pagamento,
                         motoboys=motoboys)

@app.route('/admin/despesa/<int:despesa_id>/editar-detalhes', methods=['GET', 'POST'])
@admin_required
def admin_editar_despesa_detalhes(despesa_id):
    """Editar despesa completa"""
    despesa = db.session.get(Despesa, despesa_id)
    if not despesa:
        flash('Despesa nÃ£o encontrada!', 'danger')
        return redirect(url_for('admin_caixas'))
    
    if request.method == 'POST':
        try:
            despesa.tipo = request.form.get('tipo', despesa.tipo)
            despesa.descricao = request.form.get('descricao', despesa.descricao)
            despesa.valor = parse_moeda(request.form.get('valor', despesa.valor))
            despesa.categoria_id = int(request.form.get('categoria_id')) if request.form.get('categoria_id') else None
            despesa.forma_pagamento_id = int(request.form.get('forma_pagamento_id')) if request.form.get('forma_pagamento_id') else None
            
            db.session.commit()
            flash('Despesa atualizada com sucesso!', 'success')
            return redirect(url_for('admin_visualizar_caixa', caixa_id=despesa.caixa_id))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar despesa: {str(e)}', 'danger')
    
    categorias = CategoriaDespesa.query.all()
    formas_pagamento = FormaPagamento.query.all()
    
    return render_template('admin_editar_despesa_detalhes.html',
                         despesa=despesa,
                         categorias=categorias,
                         formas_pagamento=formas_pagamento)

@app.route('/admin/delivery/<int:delivery_id>/deletar', methods=['POST'])
@admin_required
def admin_deletar_delivery(delivery_id):
    """Deletar um delivery"""
    try:
        delivery = db.session.get(Delivery, delivery_id)
        if not delivery:
            flash('Delivery nÃ£o encontrado!', 'danger')
            return redirect(url_for('admin_caixas'))
        
        caixa_id = delivery.caixa_id
        
        # Deletar pagamentos relacionados
        PagamentoDelivery.query.filter_by(delivery_id=delivery_id).delete()
        
        # Deletar delivery
        db.session.delete(delivery)
        db.session.commit()
        
        flash('Delivery removido com sucesso!', 'success')
        return redirect(url_for('admin_visualizar_caixa', caixa_id=caixa_id))
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao remover delivery: {str(e)}', 'danger')
        return redirect(url_for('admin_caixas'))

# ==================== ROUTES - RELATÃ“RIOS ====================

@app.route('/relatorios')
@login_required
def relatorios():
    usuario = db.session.get(Usuario, session['user_id'])
    if usuario.perfil in ['ADMIN', 'MASTER']:
        caixas = Caixa.query.order_by(Caixa.data.desc()).all()
    else:
        caixas = Caixa.query.filter_by(
            operador_id=usuario.id
        ).order_by(Caixa.data.desc()).all()

    return render_template(
        'relatorios.html',
        caixas=caixas
    )

@app.route('/relatorios/diario', methods=['GET', 'POST'])
@login_required
def relatorio_diario():
    """RelatÃ³rio consolidado do dia"""
    if request.method == 'POST':
        data = request.form.get('data')
        data_obj = datetime.strptime(data, '%Y-%m-%d').date()
    else:
        data_obj = datetime.now().date()
        data = data_obj.strftime('%Y-%m-%d')
    
    # Buscar todos os caixas do dia
    caixas_dia = Caixa.query.filter_by(data=data_obj).all()
    
    # Consolidar dados
    relatorio = {
        'data': data_obj,
        'total_caixas': len(caixas_dia),
        'caixas_abertos': sum(1 for c in caixas_dia if c.status == 'ABERTO'),
        'caixas_fechados': sum(1 for c in caixas_dia if c.status == 'FECHADO'),
        'total_vendas': 0,
        'total_despesas': 0,
        'total_sangrias': 0,
        'saldo_dia': 0,
        'caixas': [],  # Lista de caixas com seus totais
        'turnos': {}
    }
    
    for caixa in caixas_dia:
        totais = calcular_totais_caixa(caixa)
        relatorio['total_vendas'] += totais['total_vendas']
        relatorio['total_despesas'] += totais['despesas']
        relatorio['total_sangrias'] += sum(s.valor for s in caixa.sangrias)
        
        # Adicionar caixa com totais
        relatorio['caixas'].append({
            'caixa': caixa,
            'totais': totais
        })
        
        relatorio['turnos'][caixa.turno] = {
            'caixa': caixa,
            'totais': totais,
            'status': caixa.status
        }
    
    relatorio['saldo_dia'] = relatorio['total_vendas'] - relatorio['total_despesas'] - relatorio['total_sangrias']
    
    return render_template('relatorio_diario.html', relatorio=relatorio, data=data)

@app.route('/relatorios/turno/<int:caixa_id>')
@login_required
def relatorio_turno(caixa_id):
    """RelatÃ³rio detalhado de um turno especÃ­fico"""
    caixa = db.session.get(Caixa, caixa_id)
    if not caixa:
        flash('Caixa nÃ£o encontrado!', 'danger')
        return redirect(url_for('relatorios'))
    
    totais = calcular_totais_fechamento(caixa)
    
    return render_template('relatorio_turno.html', caixa=caixa, totais=totais)

# ==================== ROUTES - FECHAR CAIXA ====================

@app.route('/fechar-caixa')
@login_required
def fechar_caixa():
    caixa = db.session.get(Caixa, session['caixa_id'])
    if not caixa:
        session.clear()
        flash('Caixa nÃ£o encontrado. Por favor, faÃ§a login novamente.', 'warning')
        return redirect(url_for('login'))
    
    # Calcular todos os totais
    totais = calcular_totais_fechamento(caixa)
    
    return render_template('fechar_caixa.html', caixa=caixa, totais=totais)

@app.route('/fechar-caixa/confirmar', methods=['POST'])
@login_required
def confirmar_fechamento():
    try:
        caixa = db.session.get(Caixa, session['caixa_id'])
        
        # Calcular saldo final
        totais = calcular_totais_fechamento(caixa)
        caixa.saldo_final = totais['saldo_final']
        caixa.status = 'FECHADO'
        caixa.hora_fechamento = datetime.utcnow()
        
        db.session.commit()
        
        # Limpar sessÃ£o
        session.clear()
        
        flash('Caixa fechado com sucesso!', 'success')
        return redirect(url_for('login'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao fechar caixa: {str(e)}', 'danger')
        return redirect(url_for('fechar_caixa'))

# ==================== ROTAS SUPRIMENTOS (v3.0) ====================

@app.route('/suprimentos')
@login_required
def suprimentos():
    """PÃ¡gina de gestÃ£o de suprimentos"""
    caixa = db.session.get(Caixa, session.get('caixa_id'))
    if not caixa:
        flash('Caixa nÃ£o encontrado!', 'danger')
        return redirect(url_for('login'))
    
    suprimentos_list = Suprimento.query.filter_by(caixa_id=session['caixa_id']).order_by(Suprimento.data_hora.desc()).all()
    totais = calcular_totais_caixa(caixa)
    
    return render_template('suprimentos.html', suprimentos=suprimentos_list, totais=totais, caixa=caixa)

@app.route('/suprimento/novo', methods=['POST'])
@login_required
def novo_suprimento():
    """Adicionar novo suprimento"""
    try:
        valor = parse_moeda(request.form.get('valor'))
        motivo = request.form.get('motivo')
        observacao = request.form.get('observacao', '')
        
        suprimento = Suprimento(
            caixa_id=session['caixa_id'],
            valor=valor,
            motivo=motivo,
            observacao=observacao
        )
        db.session.add(suprimento)
        db.session.commit()
        
        flash(f'OK: Suprimento de R$ {valor:.2f} registrado com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao registrar suprimento: {str(e)}', 'danger')
    
    return redirect(url_for('suprimentos'))

@app.route('/admin/suprimento/<int:suprimento_id>/editar', methods=['GET', 'POST'])
@admin_required
def admin_editar_suprimento(suprimento_id):
    """Editar suprimento (somente admin)"""
    suprimento = db.session.get(Suprimento, suprimento_id)
    if not suprimento:
        flash('Suprimento nÃ£o encontrado!', 'danger')
        return redirect(url_for('suprimentos'))
    
    if request.method == 'POST':
        try:
            suprimento.valor = parse_moeda(request.form.get('valor'))
            suprimento.motivo = request.form.get('motivo')
            suprimento.observacao = request.form.get('observacao', '')
            db.session.commit()
            flash('OK: Suprimento atualizado com sucesso!', 'success')
            return redirect(url_for('suprimentos'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar: {str(e)}', 'danger')
    
    return render_template('admin_editar_suprimento.html', suprimento=suprimento)

@app.route('/admin/suprimento/<int:suprimento_id>/deletar', methods=['POST'])
@admin_required
def admin_deletar_suprimento(suprimento_id):
    """Deletar suprimento (somente admin)"""
    try:
        suprimento = db.session.get(Suprimento, suprimento_id)
        if suprimento:
            db.session.delete(suprimento)
            db.session.commit()
            flash('OK: Suprimento removido com sucesso!', 'success')
        else:
            flash('Suprimento nÃ£o encontrado!', 'warning')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao deletar: {str(e)}', 'danger')
    
    return redirect(url_for('suprimentos'))

# ==================== ROTAS EDIÃ‡ÃƒO/EXCLUSÃƒO ADMIN (v3.0) ====================

@app.route('/admin/sangria/<int:sangria_id>/editar', methods=['GET', 'POST'])
@admin_required
def admin_editar_sangria(sangria_id):
    """Editar sangria (somente admin)"""
    sangria = db.session.get(Sangria, sangria_id)
    if not sangria:
        flash('Sangria nÃ£o encontrada!', 'danger')
        return redirect(url_for('sangria'))
    
    if request.method == 'POST':
        try:
            sangria.valor = parse_moeda(request.form.get('valor'))
            sangria.motivo = request.form.get('motivo')
            sangria.observacao = request.form.get('observacao', '')
            db.session.commit()
            flash('OK: Sangria atualizada com sucesso!', 'success')
            return redirect(url_for('sangria'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar: {str(e)}', 'danger')
    
    return render_template('admin_editar_sangria.html', sangria=sangria)

@app.route('/admin/sangria/<int:sangria_id>/deletar', methods=['POST'])
@admin_required
def admin_deletar_sangria(sangria_id):
    """Deletar sangria (somente admin)"""
    try:
        sangria = db.session.get(Sangria, sangria_id)
        if sangria:
            db.session.delete(sangria)
            db.session.commit()
            flash('OK: Sangria removida com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao deletar: {str(e)}', 'danger')
    
    return redirect(url_for('sangria'))

@app.route('/admin/venda/<int:venda_id>/deletar', methods=['POST'])
@admin_required
def admin_deletar_venda_completa(venda_id):
    """Deletar venda completa (somente admin)"""
    try:
        venda = db.session.get(Venda, venda_id)
        if venda:
            # Deletar pagamentos primeiro
            PagamentoVenda.query.filter_by(venda_id=venda_id).delete()
            db.session.delete(venda)
            db.session.commit()
            flash('OK: Venda removida com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao deletar venda: {str(e)}', 'danger')
    
    return redirect(url_for('vendas'))

@app.route('/admin/delivery/<int:delivery_id>/deletar', methods=['POST'])
@admin_required
def admin_deletar_delivery_completo(delivery_id):
    """Deletar delivery completo (somente admin)"""
    try:
        delivery = db.session.get(Delivery, delivery_id)
        if delivery:
            # Deletar pagamentos primeiro
            PagamentoDelivery.query.filter_by(delivery_id=delivery_id).delete()
            db.session.delete(delivery)
            db.session.commit()
            flash('OK: Delivery removido com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao deletar delivery: {str(e)}', 'danger')
    
    return redirect(url_for('delivery'))

@app.route('/admin/despesa/<int:despesa_id>/deletar', methods=['POST'])
@admin_required
def admin_deletar_despesa_completa(despesa_id):
    """Deletar despesa completa (somente admin)"""
    try:
        despesa = db.session.get(Despesa, despesa_id)
        if despesa:
            db.session.delete(despesa)
            db.session.commit()
            flash('OK: Despesa removida com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao deletar despesa: {str(e)}', 'danger')
    
    return redirect(url_for('despesas'))

# ==================== ROTAS GESTÃƒO DE USUÃRIOS (v3.0) ====================

@app.route('/admin/usuarios/editar/<int:usuario_id>', methods=['GET', 'POST'])
@admin_required
def admin_editar_usuario(usuario_id):
    usuario = db.session.get(Usuario, usuario_id)
    if not usuario:
        flash('UsuÃ¡rio nÃ£o encontrado', 'danger')
        return redirect(url_for('configuracoes'))
    if usuario.perfil == 'MASTER':
        flash('N??o ?? permitido editar o ADMIN MASTER.', 'warning')
        return redirect(url_for('configuracoes'))

    if request.method == 'POST':
        usuario.perfil = request.form.get('perfil')
        
        # Para checkboxes, usamos 'in' para verificar se foram enviados
        usuario.acesso_dashboard = 'acesso_dashboard' in request.form
        usuario.acesso_configuracoes = 'acesso_configuracoes' in request.form
        usuario.acesso_relatorios = 'acesso_relatorios' in request.form
        
        db.session.commit()
        flash('UsuÃ¡rio atualizado com sucesso', 'success')
        return redirect(url_for('configuracoes'))

    return render_template(
        'editar_usuario.html',
        usuario=usuario
    )

@app.route('/admin/usuarios/toggle/<int:usuario_id>')
@admin_required
def admin_toggle_usuario(usuario_id):
    usuario = db.session.get(Usuario, usuario_id)
    if usuario and usuario.perfil == 'MASTER':
        flash('N??o ?? permitido alterar o ADMIN MASTER.', 'warning')
        return redirect(url_for('configuracoes'))
    if usuario and usuario.perfil not in ['ADMIN', 'MASTER']:
        usuario.ativo = not usuario.ativo
        db.session.commit()
        flash('Status do usuÃ¡rio alterado', 'info')
    return redirect(url_for('configuracoes'))

@app.route('/admin/usuario/<int:usuario_id>/editar-senha', methods=['POST'])
@admin_required
def admin_editar_senha_usuario(usuario_id):
    """Alterar senha de usuÃ¡rio (somente admin)"""
    try:
        usuario = db.session.get(Usuario, usuario_id)
        if not usuario:
            flash('UsuÃ¡rio nÃ£o encontrado!', 'danger')
            return redirect(url_for('configuracoes'))
        if usuario.perfil == 'MASTER':
            flash('N??o ?? permitido alterar a senha do ADMIN MASTER.', 'warning')
            return redirect(url_for('configuracoes'))
        
        nova_senha = request.form.get('nova_senha')
        if nova_senha and len(nova_senha) >= 3:
            usuario.senha = generate_password_hash(nova_senha)
            db.session.commit()
            flash(f'OK: Senha de {usuario.nome} alterada com sucesso!', 'success')
        else:
            flash('Senha deve ter no mínimo 3 caracteres!', 'warning')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao alterar senha: {str(e)}', 'danger')
    
    return redirect(url_for('configuracoes'))

@app.route('/admin/usuario/<int:usuario_id>/deletar', methods=['POST'])
@admin_required
def admin_deletar_usuario(usuario_id):
    """Deletar usuÃ¡rio (somente admin)"""
    try:
        usuario = db.session.get(Usuario, usuario_id)
        if not usuario:
            flash('UsuÃ¡rio nÃ£o encontrado!', 'danger')
            return redirect(url_for('configuracoes'))
        if usuario.perfil == 'MASTER':
            flash('N??o ?? permitido excluir o ADMIN MASTER.', 'warning')
            return redirect(url_for('configuracoes'))
        
        # Verificar se nÃ£o Ã© o Ãºltimo admin
        if usuario.acesso_configuracoes:
            total_admins = Usuario.query.filter_by(acesso_configuracoes=True, ativo=True).count()
            if total_admins <= 1:
                flash('Erro: Não é possível excluir o último administrador do sistema!', 'danger')
                return redirect(url_for('configuracoes'))
        
        # Verificar se tem caixas vinculados
        if len(usuario.caixas) > 0:
            flash('Erro: Não é possível excluir usuário com caixas registrados! Primeiro transfira ou remova os caixas.', 'danger')
            return redirect(url_for('configuracoes'))
        
        nome_usuario = usuario.nome
        db.session.delete(usuario)
        db.session.commit()
        flash(f'OK: Usuário {nome_usuario} excluído com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao deletar usuário: {str(e)}', 'danger')
    
    return redirect(url_for('configuracoes'))

# ==================== ROTAS PARA FORMAS DE PAGAMENTO ====================

@app.route('/admin/forma-pagamento/<int:forma_id>/toggle', methods=['POST'])
@admin_required
def admin_toggle_forma_pagamento(forma_id):
    """Ativar/Desativar forma de pagamento"""
    forma = db.session.get(FormaPagamento, forma_id)
    if forma:
        forma.ativo = not forma.ativo
        db.session.commit()
        flash(f'Forma de pagamento "{forma.nome}" atualizada!', 'success')
    return redirect(url_for('configuracoes'))

@app.route('/admin/forma-pagamento/<int:forma_id>/editar', methods=['POST'])
@admin_required
def admin_editar_forma_pagamento(forma_id):
    """Editar forma de pagamento"""
    try:
        forma = db.session.get(FormaPagamento, forma_id)
        if forma:
            forma.nome = request.form.get('nome')
            forma.ativo = 'ativo' in request.form
            db.session.commit()
            flash('Forma de pagamento atualizada!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao atualizar: {str(e)}', 'danger')
    return redirect(url_for('configuracoes'))

@app.route('/admin/forma-pagamento/<int:forma_id>/deletar', methods=['POST'])
@admin_required
def admin_deletar_forma_pagamento(forma_id):
    """Deletar forma de pagamento"""
    try:
        forma = db.session.get(FormaPagamento, forma_id)
        if forma:
            db.session.delete(forma)
            db.session.commit()
            flash('Forma de pagamento excluÃ­da!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir: {str(e)}', 'danger')
    return redirect(url_for('configuracoes'))

# ==================== ROTAS PARA BANDEIRAS ====================

@app.route('/admin/bandeira/<int:bandeira_id>/toggle', methods=['POST'])
@admin_required
def admin_toggle_bandeira(bandeira_id):
    """Ativar/Desativar bandeira"""
    bandeira = db.session.get(BandeiraCartao, bandeira_id)
    if bandeira:
        bandeira.ativo = not bandeira.ativo
        db.session.commit()
        flash(f'Bandeira "{bandeira.nome}" atualizada!', 'success')
    return redirect(url_for('configuracoes'))

@app.route('/admin/bandeira/<int:bandeira_id>/editar', methods=['POST'])
@admin_required
def admin_editar_bandeira(bandeira_id):
    """Editar bandeira"""
    try:
        bandeira = db.session.get(BandeiraCartao, bandeira_id)
        if bandeira:
            bandeira.nome = request.form.get('nome')
            bandeira.ativo = 'ativo' in request.form
            db.session.commit()
            flash('Bandeira atualizada!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao atualizar: {str(e)}', 'danger')
    return redirect(url_for('configuracoes'))

@app.route('/admin/bandeira/<int:bandeira_id>/deletar', methods=['POST'])
@admin_required
def admin_deletar_bandeira(bandeira_id):
    """Deletar bandeira"""
    try:
        bandeira = db.session.get(BandeiraCartao, bandeira_id)
        if bandeira:
            db.session.delete(bandeira)
            db.session.commit()
            flash('Bandeira excluÃ­da!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir: {str(e)}', 'danger')
    return redirect(url_for('configuracoes'))

# ==================== ROTAS PARA CATEGORIAS ====================

@app.route('/admin/categoria/<int:categoria_id>/toggle', methods=['POST'])
@admin_required
def admin_toggle_categoria(categoria_id):
    """Ativar/Desativar categoria"""
    categoria = db.session.get(CategoriaDespesa, categoria_id)
    if categoria:
        categoria.ativo = not categoria.ativo
        db.session.commit()
        flash(f'Categoria "{categoria.nome}" atualizada!', 'success')
    return redirect(url_for('configuracoes'))

@app.route('/admin/categoria/<int:categoria_id>/editar', methods=['POST'])
@admin_required
def admin_editar_categoria(categoria_id):
    """Editar categoria"""
    try:
        categoria = db.session.get(CategoriaDespesa, categoria_id)
        if categoria:
            categoria.nome = request.form.get('nome')
            categoria.tipo = request.form.get('tipo')
            categoria.ativo = 'ativo' in request.form
            db.session.commit()
            flash('Categoria atualizada!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao atualizar: {str(e)}', 'danger')
    return redirect(url_for('configuracoes'))

@app.route('/admin/categoria/<int:categoria_id>/deletar', methods=['POST'])
@admin_required
def admin_deletar_categoria(categoria_id):
    """Deletar categoria"""
    try:
        categoria = db.session.get(CategoriaDespesa, categoria_id)
        if categoria:
            db.session.delete(categoria)
            db.session.commit()
            flash('Categoria excluÃ­da!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir: {str(e)}', 'danger')
    return redirect(url_for('configuracoes'))

# ==================== ROTAS PARA MOTOBOYS ====================

@app.route('/admin/motoboy/<int:motoboy_id>/toggle', methods=['POST'])
@admin_required
def admin_toggle_motoboy(motoboy_id):
    """Ativar/Desativar motoboy"""
    motoboy = db.session.get(Motoboy, motoboy_id)
    if motoboy:
        motoboy.ativo = not motoboy.ativo
        db.session.commit()
        flash(f'Motoboy "{motoboy.nome}" atualizado!', 'success')
    return redirect(url_for('configuracoes'))

@app.route('/admin/motoboy/<int:motoboy_id>/editar', methods=['POST'])
@admin_required
def admin_editar_motoboy(motoboy_id):
    """Editar motoboy"""
    try:
        motoboy = db.session.get(Motoboy, motoboy_id)
        if motoboy:
            motoboy.nome = request.form.get('nome')
            motoboy.taxa_padrao = parse_moeda(request.form.get('taxa_padrao', 5.00))
            motoboy.ativo = 'ativo' in request.form
            db.session.commit()
            flash('Motoboy atualizado!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao atualizar: {str(e)}', 'danger')
    return redirect(url_for('configuracoes'))

@app.route('/admin/motoboy/<int:motoboy_id>/deletar', methods=['POST'])
@admin_required
def admin_deletar_motoboy(motoboy_id):
    """Deletar motoboy"""
    try:
        motoboy = db.session.get(Motoboy, motoboy_id)
        if motoboy:
            db.session.delete(motoboy)
            db.session.commit()
            flash('Motoboy excluÃ­do!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir: {str(e)}', 'danger')
    return redirect(url_for('configuracoes'))

@app.route('/admin/licenca/<int:licenca_id>/toggle', methods=['POST'])
def toggle_licenca(licenca_id):
    """Ativar/Desativar licenÃ§a"""
    licenca = db.session.get(Licenca, licenca_id)
    if licenca:
        licenca.ativo = not licenca.ativo
        db.session.commit()
        flash(f'LicenÃ§a {licenca.email} {"ativada" if licenca.ativo else "desativada"}!', 'success')
    return redirect(url_for('admin_licencas'))

@app.route('/admin/licenca/<int:licenca_id>/deletar', methods=['POST'])
def deletar_licenca(licenca_id):
    """Deletar licenÃ§a"""
    try:
        licenca = db.session.get(Licenca, licenca_id)
        if licenca:
            # Deletar dispositivos associados primeiro
            Dispositivo.query.filter_by(licenca_id=licenca_id).delete()
            # Deletar licenÃ§a
            db.session.delete(licenca)
            db.session.commit()
            flash(f'LicenÃ§a {licenca.email} excluÃ­da com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir licenÃ§a: {str(e)}', 'danger')
    return redirect(url_for('admin_licencas'))



#+++++++++++++++++++++teste
@app.route('/debug/licencas')
def debug_licencas():
    """Debug: Mostrar licenÃ§as no terminal"""
    licencas = Licenca.query.all()
    print("\n" + "="*50)
    print("LICENÃ‡AS CADASTRADAS NO SISTEMA:")
    print("="*50)
    
    for licenca in licencas:
        print(f"\nID: {licenca.id}")
        print(f"E-mail: {licenca.email}")
        print(f"Chave: (oculta)")
        print(f"Status: {licenca.status}")
        print(f"AtivaÃ§Ã£o: {licenca.data_ativacao}")
        print(f"ExpiraÃ§Ã£o: {licenca.data_expiracao}")
        print(f"Ativa: {'SIM' if licenca.ativo else 'NÃƒO'}")
        print(f"Dispositivos: {len(licenca.dispositivos)}/{licenca.max_dispositivos}")
    
    print("\n" + "="*50)
    return "Verifique o terminal do Python para ver as licenÃ§as cadastradas."



# ==================== ROTA EXPORTAR TODOS CAIXAS ====================

@app.route('/exportar/todos-caixas')
@admin_required
def exportar_todos_caixas():
    """Exportar todos os caixas para CSV"""
    import csv
    from io import StringIO
    from flask import Response
    
    output = StringIO()
    writer = csv.writer(output, delimiter=',')

    
    # CabeÃ§alho
    writer.writerow([
        'ID', 'Data', 'Turno', 'Operador', 'Status', 'Saldo Inicial', 'Saldo Final',
        'Total Vendas', 'Total Despesas', 'Total Sangrias', 'Hora Abertura', 'Hora Fechamento'
    ])
    
    # Buscar todos os caixas
    caixas = Caixa.query.order_by(Caixa.data.desc(), Caixa.turno).all()
    
    for caixa in caixas:
        totais = calcular_totais_caixa(caixa)
        
        writer.writerow([
            caixa.id,
            caixa.data.strftime('%d/%m/%Y'),
            caixa.turno,
            caixa.operador.nome,
            caixa.status,
            f'{caixa.saldo_inicial:.2f}',
            f'{caixa.saldo_final:.2f}' if caixa.saldo_final else '',
            f'{totais["total_vendas"]:.2f}',
            f'{totais["despesas"]:.2f}',
            f'{totais["sangrias"]:.2f}',
            caixa.hora_abertura.strftime('%H:%M:%S'),
            caixa.hora_fechamento.strftime('%H:%M:%S') if caixa.hora_fechamento else ''
        ])
    
    output.seek(0)
    return Response(
        output.getvalue().encode('utf-8-sig'),
        mimetype='text/csv; charset=utf-8',
        headers={
            'Content-Disposition': 'attachment; filename=todos_caixas.csv'
        }
    )

# ==================== ROTA EXPORTAR EXCEL (v3.0) ====================

@app.route('/exportar/excel/<int:caixa_id>')
@login_required
def exportar_excel_caixa(caixa_id):
    """Exportar todos os movimentos do caixa para Excel/CSV"""
    import csv
    from io import StringIO
    from flask import Response
    
    caixa = db.session.get(Caixa, caixa_id)
    if not caixa:
        flash('Caixa nÃ£o encontrado!', 'danger')
        return redirect(url_for('dashboard'))
    
    # Verificar permissÃ£o (operador sÃ³ pode exportar prÃ³prio caixa)
    usuario = db.session.get(Usuario, session['user_id'])
    if not usuario.acesso_configuracoes and caixa.operador_id != usuario.id:
        flash('VocÃª nÃ£o tem permissÃ£o para exportar este caixa!', 'danger')
        return redirect(url_for('dashboard'))
    
    output = StringIO()
    writer = csv.writer(output, delimiter=',')
    
    # CabeÃ§alho com TODAS as colunas solicitadas
    writer.writerow([
        'ID',
        'Data',
        'Hora',
        'Turno',
        'Operador',
        'Tipo Movimento',
        'Tipo Venda',
        'NÃºmero Mesa/BalcÃ£o',
        'Cliente',
        'EndereÃ§o',
        'Telefone',
        'Valor Bruto',
        'Valor LÃ­quido',
        'Forma Pagamento',
        'Bandeira',
        'Taxa Entrega',
        'Motoboy',
        'Categoria Despesa',
        'DescriÃ§Ã£o',
        'ObservaÃ§Ãµes',
        'Nota Fiscal'
    ])
    
    # VENDAS
    for venda in caixa.vendas:
        for pag in venda.pagamentos:
            writer.writerow([
                venda.id,
                venda.data_hora.strftime('%d/%m/%Y'),
                venda.data_hora.strftime('%H:%M:%S'),
                caixa.turno,
                caixa.operador.nome,
                'VENDA',
                venda.tipo,
                venda.numero if venda.numero else '-',
                '-',
                '-',
                '-',
                f'{pag.valor:.2f}',
                f'{pag.valor:.2f}',
                pag.forma_pagamento.nome if pag.forma_pagamento else '-',
                pag.bandeira.nome if pag.bandeira else '-',
                '-',
                '-',
                '-',
                '-',
                venda.observacao if venda.observacao else '-',
                'Sim' if venda.emitiu_nota else 'NÃ£o'
            ])
    
    # DELIVERIES
    for delivery in caixa.deliveries:
        total_bruto = delivery.total + delivery.taxa_entrega
        for pag in delivery.pagamentos:
            bandeira_nome = '-'
            if hasattr(pag, 'bandeira') and pag.bandeira:
                bandeira_nome = pag.bandeira.nome
            
            writer.writerow([
                delivery.id,
                delivery.data_hora.strftime('%d/%m/%Y'),
                delivery.data_hora.strftime('%H:%M:%S'),
                caixa.turno,
                caixa.operador.nome,
                'DELIVERY',
                '-',
                '-',
                delivery.cliente,
                getattr(delivery, 'endereco', '-') or '-',
                getattr(delivery, 'telefone', '-') or '-',
                f'{total_bruto:.2f}',
                f'{pag.valor:.2f}',
                pag.forma_pagamento.nome if pag.forma_pagamento else '-',
                bandeira_nome,
                f'{delivery.taxa_entrega:.2f}',
                delivery.motoboy.nome if delivery.motoboy else '-',
                '-',
                '-',
                delivery.observacao if delivery.observacao else '-',
                'Sim' if delivery.emitiu_nota else 'NÃ£o'
            ])
    
    # DESPESAS
    for despesa in caixa.despesas:
        writer.writerow([
            despesa.id,
            despesa.data_hora.strftime('%d/%m/%Y'),
            despesa.data_hora.strftime('%H:%M:%S'),
            caixa.turno,
            caixa.operador.nome,
            'DESPESA',
            despesa.tipo,
            '-',
            '-',
            '-',
            '-',
            f'{despesa.valor:.2f}',
            f'{despesa.valor:.2f}',
            despesa.forma_pagamento.nome if despesa.forma_pagamento else '-',
            '-',
            '-',
            '-',
            despesa.categoria.nome if despesa.categoria else '-',
            despesa.descricao,
            despesa.observacao if despesa.observacao else '-',
            '-'
        ])
    
    # SANGRIAS
    for sangria in caixa.sangrias:
        writer.writerow([
            sangria.id,
            sangria.data_hora.strftime('%d/%m/%Y'),
            sangria.data_hora.strftime('%H:%M:%S'),
            caixa.turno,
            caixa.operador.nome,
            'SANGRIA',
            '-',
            '-',
            '-',
            '-',
            '-',
            f'{sangria.valor:.2f}',
            f'-{sangria.valor:.2f}',
            '-',
            '-',
            '-',
            '-',
            '-',
            sangria.motivo,
            sangria.observacao if sangria.observacao else '-',
            '-'
        ])
    
    # SUPRIMENTOS
    try:
        if hasattr(caixa, 'suprimentos'):
            for suprimento in caixa.suprimentos:
                writer.writerow([
                    suprimento.id,
                    suprimento.data_hora.strftime('%d/%m/%Y'),
                    suprimento.data_hora.strftime('%H:%M:%S'),
                    caixa.turno,
                    caixa.operador.nome,
                    'SUPRIMENTO',
                    '-',
                    '-',
                    '-',
                    '-',
                    '-',
                    f'{suprimento.valor:.2f}',
                    f'{suprimento.valor:.2f}',
                    '-',
                    '-',
                    '-',
                    '-',
                    '-',
                    suprimento.motivo,
                    suprimento.observacao if suprimento.observacao else '-',
                    '-'
                ])
    except:
        pass  # Caso tabela suprimento nÃ£o exista ainda
    
    output.seek(0)
    return Response(
        output.getvalue().encode('utf-8-sig'),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': 'attachment; filename=todos_caixas.csv'}
    )

# ==================== HELPER FUNCTIONS ====================

def calcular_totais_caixa(caixa):
    """Calcula todos os totais do caixa"""
    totais = {
        'vendas_loja': 0,
        'vendas_delivery': 0,
        'total_vendas': 0,
        'dinheiro': 0,
        'credito': 0,
        'debito': 0,
        'pix': 0,
        'online': 0,
        'notas_fiscais': 0,
        'despesas': 0,
        'sangrias': 0,
        'saldo_atual': caixa.saldo_inicial
    }
    
    # Vendas
    for venda in caixa.vendas:
        totais['vendas_loja'] += venda.total
        if venda.emitiu_nota:
            totais['notas_fiscais'] += venda.total
        
        for pagamento in venda.pagamentos:
            if not pagamento.forma_pagamento:
                # Forma de pagamento removida do cadastro
                continue
            forma = pagamento.forma_pagamento.nome.upper()
            if 'DINHEIRO' in forma:
                totais['dinheiro'] += pagamento.valor
            elif 'CRÃ‰DITO' in forma or 'CREDITO' in forma:
                totais['credito'] += pagamento.valor
            elif 'DÃ‰BITO' in forma or 'DEBITO' in forma:
                totais['debito'] += pagamento.valor
            elif 'PIX' in forma:
                totais['pix'] += pagamento.valor
            elif 'ONLINE' in forma:
                totais['online'] += pagamento.valor
    
    # Delivery
    for delivery in caixa.deliveries:
        totais['vendas_delivery'] += delivery.total + delivery.taxa_entrega
        if delivery.emitiu_nota:
            totais['notas_fiscais'] += delivery.total + delivery.taxa_entrega
        
        for pagamento in delivery.pagamentos:
            if not pagamento.forma_pagamento:
                # Forma de pagamento removida do cadastro
                continue
            forma = pagamento.forma_pagamento.nome.upper()
            if 'DINHEIRO' in forma:
                totais['dinheiro'] += pagamento.valor
            elif 'CRÃ‰DITO' in forma or 'CREDITO' in forma:
                totais['credito'] += pagamento.valor
            elif 'DÃ‰BITO' in forma or 'DEBITO' in forma:
                totais['debito'] += pagamento.valor
            elif 'PIX' in forma:
                totais['pix'] += pagamento.valor
            elif 'ONLINE' in forma:
                totais['online'] += pagamento.valor
    
    # Despesas
    for despesa in caixa.despesas:
        totais['despesas'] += despesa.valor
    
    # Sangrias
    for sangria in caixa.sangrias:
        totais['sangrias'] += sangria.valor
    
    totais['total_vendas'] = totais['vendas_loja'] + totais['vendas_delivery']
    totais['saldo_atual'] = caixa.saldo_inicial + totais['total_vendas'] + totais.get('suprimentos', 0) - totais['despesas'] - totais['sangrias']
    
    return totais

def calcular_totais_delivery(caixa):
    """Calcula totais especÃ­ficos do delivery"""
    totais = {
        'total_delivery': 0,
        'total_taxas': 0,
        'quantidade_pedidos': 0,
        'motoboys': {}
    }
    
    for delivery in caixa.deliveries:
        totais['total_delivery'] += delivery.total + delivery.taxa_entrega
        totais['total_taxas'] += delivery.taxa_entrega
        totais['quantidade_pedidos'] += 1
        
        if delivery.motoboy:
            nome = delivery.motoboy.nome
            if nome not in totais['motoboys']:
                totais['motoboys'][nome] = 0
            totais['motoboys'][nome] += delivery.taxa_entrega
    
    return totais

def calcular_totais_fechamento(caixa):
    """Calcula todos os totais para fechamento"""
    totais = calcular_totais_caixa(caixa)
    
    # Adicionar informaÃ§Ãµes extras para fechamento
    totais['despesas_fixas'] = sum(d.valor for d in caixa.despesas if d.tipo == 'FIXA')
    totais['despesas_variaveis'] = sum(d.valor for d in caixa.despesas if d.tipo == 'VARIAVEL')
    totais['despesas_saidas'] = sum(d.valor for d in caixa.despesas if d.tipo == 'SAIDA')
    totais['saldo_final'] = totais['saldo_atual']
    
    return totais

def calcular_metricas_dashboard(caixas):
    """Calcula mÃ©tricas para o dashboard - CORRIGIDO"""
    metricas = {
        'total_receitas': 0,
        'total_despesas': 0,
        'saldo_liquido': 0,
        'ticket_medio': 0,
        'total_transacoes': 0,
        'formas_pagamento': {},
        'tipos_venda': {'MESA': 0, 'BALCAO': 0, 'DELIVERY': 0},
        'despesas_categoria': {},
        'vendas_count': 0,
        'delivery_count': 0
    }
    
    for caixa in caixas:
        totais = calcular_totais_caixa(caixa)
        metricas['total_receitas'] += totais.get('total_vendas', 0)
        metricas['total_despesas'] += totais.get('despesas', 0)
        
        # Contar transaÃ§Ãµes
        metricas['vendas_count'] += len(caixa.vendas)
        metricas['delivery_count'] += len(caixa.deliveries)
        metricas['total_transacoes'] = metricas['vendas_count'] + metricas['delivery_count']
        
        # Formas de pagamento
        # Vendas
        for venda in caixa.vendas:
            for pagamento in venda.pagamentos:
                if not pagamento.forma_pagamento:
                    continue
                forma = pagamento.forma_pagamento.nome
                metricas['formas_pagamento'][forma] = metricas['formas_pagamento'].get(forma, 0) + pagamento.valor
            
            # Tipos de venda
            metricas['tipos_venda'][venda.tipo] += venda.total
        
        # Deliveries
        for delivery in caixa.deliveries:
            for pagamento in delivery.pagamentos:
                if not pagamento.forma_pagamento:
                    continue
                forma = pagamento.forma_pagamento.nome
                metricas['formas_pagamento'][forma] = metricas['formas_pagamento'].get(forma, 0) + pagamento.valor
            
            metricas['tipos_venda']['DELIVERY'] += delivery.total + delivery.taxa_entrega
        
        # Despesas por categoria
        for despesa in caixa.despesas:
            if despesa.categoria:
                cat = despesa.categoria.nome
                metricas['despesas_categoria'][cat] = metricas['despesas_categoria'].get(cat, 0) + despesa.valor
    
    metricas['saldo_liquido'] = metricas['total_receitas'] - metricas['total_despesas']
    metricas['ticket_medio'] = metricas['total_receitas'] / metricas['total_transacoes'] if metricas['total_transacoes'] > 0 else 0
    try:
        metricas['ticket_medio'] = round(float(metricas['ticket_medio']), 2)
    except Exception:
        metricas['ticket_medio'] = 0
    
    return metricas

def calcular_metricas_avancadas(caixas):
    """Calcula mÃ©tricas avanÃ§adas para dashboard - CORRIGIDO"""
    metricas = {
        'vendas_por_turno': {'MANHÃƒ': 0, 'TARDE': 0, 'NOITE': 0},
        'transacoes_por_turno': {'MANHÃƒ': 0, 'TARDE': 0, 'NOITE': 0},
        'motoboys_taxas': {},
        'despesas_por_tipo': {'FIXA': 0, 'VARIAVEL': 0, 'SAIDA': 0},
        'contas_assinadas': 0,
        'total_sangrias': 0,
        'margem_lucro': 0,
        'custo_operacional': 0,
        'lucratividade': 0,
        'vendas_por_dia': {},
        'despesas_por_dia': {},
        'melhor_dia': {'dia': '-', 'valor': 0},
        'pior_dia': {'dia': '-', 'valor': 99999999},
        'total_notas_fiscais': 0,
        'percentual_notas': 0,
        'ticket_medio_mesa': 0,
        'ticket_medio_delivery': 0,
        'total_produtos_vendidos': 0,
        'vendas_mesa_count': 0,
        'vendas_balcao_count': 0,
        'vendas_delivery_count': 0
    }
    metricas['vendas_por_hora'] = {f'{h:02d}h': 0 for h in range(24)}
    
    total_vendas_mesa = 0
    count_vendas_mesa = 0
    total_vendas_balcao = 0
    count_vendas_balcao = 0
    total_vendas_delivery = 0
    count_vendas_delivery = 0

    def normalizar_turno(turno):
        if not turno:
            return 'MANHÃƒ'
        t = str(turno).upper()
        if 'MANH' in t:
            return 'MANHÃƒ'
        if 'TARDE' in t:
            return 'TARDE'
        if 'NOITE' in t:
            return 'NOITE'
        return 'MANHÃƒ'
    
    for caixa in caixas:
        turno = normalizar_turno(caixa.turno)
        dia_str = caixa.data.strftime('%d/%m')
        
        # Inicializar dia se nÃ£o existe
        if dia_str not in metricas['vendas_por_dia']:
            metricas['vendas_por_dia'][dia_str] = 0
        if dia_str not in metricas['despesas_por_dia']:
            metricas['despesas_por_dia'][dia_str] = 0
        
        # Vendas
        for venda in caixa.vendas:
            metricas['vendas_por_turno'][turno] += venda.total
            metricas['transacoes_por_turno'][turno] += 1
            metricas['vendas_por_dia'][dia_str] += venda.total
            hora_venda = f"{venda.data_hora.hour:02d}h"
            metricas['vendas_por_hora'][hora_venda] += venda.total
            
            if venda.emitiu_nota:
                metricas['total_notas_fiscais'] += venda.total
            
            # Contar por tipo
            if venda.tipo == 'MESA':
                total_vendas_mesa += venda.total
                count_vendas_mesa += 1
                metricas['vendas_mesa_count'] += 1
            elif venda.tipo == 'BALCAO':
                total_vendas_balcao += venda.total
                count_vendas_balcao += 1
                metricas['vendas_balcao_count'] += 1
            
            # Contar contas assinadas
            for pagamento in venda.pagamentos:
                if not pagamento.forma_pagamento:
                    continue
                if 'ASSINADA' in pagamento.forma_pagamento.nome.upper() or 'CONTA' in pagamento.forma_pagamento.nome.upper():
                    metricas['contas_assinadas'] += pagamento.valor
        
        # Delivery
        for delivery in caixa.deliveries:
            total = delivery.total + delivery.taxa_entrega
            metricas['vendas_por_turno'][turno] += total
            metricas['transacoes_por_turno'][turno] += 1
            metricas['vendas_por_dia'][dia_str] += total
            hora_delivery = f"{delivery.data_hora.hour:02d}h"
            metricas['vendas_por_hora'][hora_delivery] += total
            metricas['vendas_delivery_count'] += 1
            
            if delivery.emitiu_nota:
                metricas['total_notas_fiscais'] += total
            
            # Ticket mÃ©dio delivery
            total_vendas_delivery += total
            count_vendas_delivery += 1
            
            # Taxas por motoboy
            if delivery.motoboy:
                nome = delivery.motoboy.nome
                if nome not in metricas['motoboys_taxas']:
                    metricas['motoboys_taxas'][nome] = {'total': 0, 'quantidade': 0}
                metricas['motoboys_taxas'][nome]['total'] += delivery.taxa_entrega
                metricas['motoboys_taxas'][nome]['quantidade'] += 1
            
            # Contar contas assinadas em delivery
            for pagamento in delivery.pagamentos:
                if not pagamento.forma_pagamento:
                    continue
                if 'ASSINADA' in pagamento.forma_pagamento.nome.upper() or 'CONTA' in pagamento.forma_pagamento.nome.upper():
                    metricas['contas_assinadas'] += pagamento.valor
        
        # Despesas por tipo
        for despesa in caixa.despesas:
            if despesa.tipo in metricas['despesas_por_tipo']:
                metricas['despesas_por_tipo'][despesa.tipo] += despesa.valor
        
        # Sangrias
        for sangria in caixa.sangrias:
            metricas['total_sangrias'] += sangria.valor
    
    # Calcular melhor e pior dia
    if metricas['vendas_por_dia']:
        for dia, valor in metricas['vendas_por_dia'].items():
            if valor > metricas['melhor_dia']['valor']:
                metricas['melhor_dia'] = {'dia': dia, 'valor': valor}
            if valor < metricas['pior_dia']['valor']:
                metricas['pior_dia'] = {'dia': dia, 'valor': valor}
    
    # Calcular tickets mÃ©dios
    metricas['ticket_medio_mesa'] = total_vendas_mesa / count_vendas_mesa if count_vendas_mesa > 0 else 0
    metricas['ticket_medio_delivery'] = total_vendas_delivery / count_vendas_delivery if count_vendas_delivery > 0 else 0
    # Ticket médio balcão (quando houver vendas no balcão)
    metricas['ticket_medio_balcao'] = total_vendas_balcao / count_vendas_balcao if count_vendas_balcao > 0 else 0
    
    # Calcular mÃ©tricas financeiras
    total_receitas = sum(metricas['vendas_por_turno'].values())
    total_despesas = sum(metricas['despesas_por_tipo'].values())
    
    metricas['custo_operacional'] = total_despesas
    metricas['margem_lucro'] = ((total_receitas - total_despesas) / total_receitas * 100) if total_receitas > 0 else 0
    metricas['lucratividade'] = total_receitas - total_despesas
    metricas['percentual_notas'] = (metricas['total_notas_fiscais'] / total_receitas * 100) if total_receitas > 0 else 0
    
    return metricas

# ==================== INITIALIZE DATABASE ====================

def init_db():
    """Inicializa o banco de dados com dados padrÃ£o"""
    with app.app_context():
        db.create_all()

        def _env_password(var_name):
            value = os.environ.get(var_name)
            value = (value or '').strip()
            return value or None

        def _generate_password():
            # URL-safe and reasonably strong for first-run bootstrap
            return secrets.token_urlsafe(12)

        def _hash_password(plain):
            return generate_password_hash(plain)
        
        # Verificar se ja existem dados
        if Usuario.query.first():
            master = Usuario.query.filter_by(nome='ADMIN MASTER').first()
            if master:
                if master.perfil != 'MASTER':
                    master.perfil = 'MASTER'
                master.acesso_dashboard = True
                master.acesso_configuracoes = True
                master.acesso_relatorios = True
                master.ativo = True
                db.session.commit()
            else:
                master = Usuario.query.filter_by(perfil='MASTER').first()
                if not master:
                    master_password = _env_password('MASTER_PASSWORD') or _generate_password()
                    master = Usuario(
                        nome='ADMIN MASTER',
                        senha=_hash_password(master_password),
                        perfil='MASTER',
                        acesso_dashboard=True,
                        acesso_configuracoes=True,
                        acesso_relatorios=True
                    )
                    db.session.add(master)
                    db.session.commit()
                    print("⚠️  Usuário 'ADMIN MASTER' recriado automaticamente (MASTER_PASSWORD não definido).")
                    print("⚠️  Altere a senha imediatamente pelo painel de usuários.")
                    print("🔐 Senha temporária criada — valor ocultado por segurança. Defina MASTER_PASSWORD no ambiente para controlar esta senha.")

            # Garantir que exista ao menos um admin (somente se não houver nenhum)
            admin_exists = Usuario.query.filter_by(perfil='ADMIN', ativo=True).first() is not None
            if not admin_exists and Usuario.query.filter_by(nome='admin').first() is None:
                admin_password = _env_password('ADMIN_PASSWORD') or _generate_password()
                admin = Usuario(
                    nome='admin',
                    senha=_hash_password(admin_password),
                    perfil='ADMIN',
                    acesso_dashboard=True,
                    acesso_configuracoes=True,
                    acesso_relatorios=True,
                    ativo=True
                )
                db.session.add(admin)
                db.session.commit()
                print("⚠️  Usuário 'admin' criado automaticamente (ADMIN_PASSWORD não definido).")
                print("⚠️  Altere a senha imediatamente pelo painel de usuários.")
                print("🔐 Senha temporária criada — valor ocultado por segurança. Defina ADMIN_PASSWORD no ambiente para controlar esta senha.")

            return
        
        # Criar usuario admin master
        master_password = _env_password('MASTER_PASSWORD') or _generate_password()
        master = Usuario(
            nome='ADMIN MASTER',
            senha=_hash_password(master_password),
            perfil='MASTER',
            acesso_dashboard=True,
            acesso_configuracoes=True,
            acesso_relatorios=True
        )
        db.session.add(master)

        # Criar usuario admin
        admin_password = _env_password('ADMIN_PASSWORD') or _generate_password()
        admin = Usuario(
            nome='admin',
            senha=_hash_password(admin_password),
            perfil='ADMIN',
            acesso_dashboard=True,
            acesso_configuracoes=True,
            acesso_relatorios=True
        )
        db.session.add(admin)

        # Exibir senhas temporárias apenas no primeiro seed
        if not _env_password('MASTER_PASSWORD'):
            print("✅ Seed inicial: usuário 'ADMIN MASTER' criado.")
            print("⚠️  Defina MASTER_PASSWORD para evitar senha aleatória.")
            print("🔐 Senha temporária criada — valor ocultado por segurança. Defina MASTER_PASSWORD no ambiente para controlar esta senha.")
        if not _env_password('ADMIN_PASSWORD'):
            print("✅ Seed inicial: usuário 'admin' criado.")
            print("⚠️  Defina ADMIN_PASSWORD para evitar senha aleatória.")
            print("🔐 Senha temporária criada — valor ocultado por segurança. Defina ADMIN_PASSWORD no ambiente para controlar esta senha.")

        # Criar formas de pagamento padrÃ£o
        formas = [
            'Dinheiro', 'CrÃ©dito', 'DÃ©bito', 'PIX', 'CartÃ£o (Voucher)',
            'Conta Assinada', 'PG Online', 'Link de Pagamento',
            'TransferÃªncia', 'DepÃ³sito', 'Boleto', 'Cheque',
            'Vale RefeiÃ§Ã£o', 'Vale AlimentaÃ§Ã£o', 'Cortesia'
        ]
        for forma in formas:
            db.session.add(FormaPagamento(nome=forma))
        
        # Criar bandeiras padrÃ£o
        bandeiras = [
            'Visa', 'Mastercard', 'Elo', 'American Express',
            'Hipercard', 'Diners Club', 'Discover', 'Aura',
            'Cabal', 'Banescard', 'Good Card', 'Sodexo',
            'Ticket', 'VR', 'Alelo'
        ]
        for bandeira in bandeiras:
            db.session.add(BandeiraCartao(nome=bandeira))
        
        # Criar categorias padrÃ£o
        categorias = [
            ('Aluguel', 'FIXA'),
            ('CondomÃ­nio', 'FIXA'),
            ('Ãgua', 'FIXA'),
            ('Luz', 'FIXA'),
            ('Internet', 'FIXA'),
            ('Telefonia', 'FIXA'),
            ('Contabilidade', 'FIXA'),
            ('Sistema/Software', 'FIXA'),
            ('Produtos', 'VARIAVEL'),
            ('Embalagens', 'VARIAVEL'),
            ('GÃ¡s', 'VARIAVEL'),
            ('ManutenÃ§Ã£o', 'VARIAVEL'),
            ('Limpeza', 'VARIAVEL'),
            ('Marketing', 'VARIAVEL'),
            ('Fretado/Entrega', 'VARIAVEL'),
            ('ComissÃ£o Motoboy', 'VARIAVEL'),
            ('Taxas CartÃ£o', 'SAIDA'),
            ('Taxas Plataforma', 'SAIDA'),
            ('Impostos', 'SAIDA'),
            ('Passagem', 'SAIDA'),
            ('Multas', 'SAIDA'),
            ('Outros', 'SAIDA')
        ]
        for nome, tipo in categorias:
            db.session.add(CategoriaDespesa(nome=nome, tipo=tipo))
        
        # Criar motoboys padrÃ£o
        motoboys = ['JoÃ£o', 'Maria', 'Pedro']
        for nome in motoboys:
            db.session.add(Motoboy(nome=nome, taxa_padrao=5.00))
        
        db.session.commit()
        print("OK: Banco de dados inicializado com sucesso!")

@app.route('/criar-chave-teste')
def criar_chave_teste():
    """Criar uma chave de teste"""
    import random
    import string
    
    # Gerar chave no formato XXXX-XXXX-XXXX-XXXX
    chave = '-'.join(
        ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
        for _ in range(4)
    )
    
    # Criar licenÃ§a de teste
    licenca = Licenca(
        email='teste@email.com',
        chave_ativacao=chave,
        data_expiracao=datetime.utcnow() + timedelta(days=365),
        status='ATIVA'
    )
    db.session.add(licenca)
    db.session.commit()
    
    return f"""
    <h1>Chave de Teste Criada!</h1>
    <h3>E-mail: teste@email.com</h3>
    <h3>Chave: <code>{chave}</code></h3>
    <button onclick="navigator.clipboard.writeText('{chave}')">Copiar Chave</button>
    <br><br>
    <a href="/ativacao">Ir para tela de ativaÃ§Ã£o</a>
    """

@app.route('/ver-chaves')
def ver_chaves():
    """Rota temporÃ¡ria para ver chaves de licenÃ§a"""
    licencas = Licenca.query.all()
    
    html = """
    <!DOCTYPE html>
    <html>
    <head>
        <title>Chaves de LicenÃ§a</title>
        <style>
            body { font-family: Arial; padding: 20px; }
            .chave { background: #f0f0f0; padding: 10px; margin: 10px; border-radius: 5px; }
            code { font-size: 16px; color: #d63384; }
        </style>
    </head>
    <body>
        <h1>Chaves de LicenÃ§a Cadastradas</h1>
    """
    
    if licencas:
        for licenca in licencas:
            html += f"""
            <div class="chave">
                <strong>ID:</strong> {licenca.id}<br>
                <strong>E-mail:</strong> {licenca.email}<br>
                <strong>Chave:</strong> <code>{licenca.chave_ativacao}</code><br>
                <strong>Status:</strong> {licenca.status}<br>
                <strong>Ativa:</strong> {licenca.ativo}<br>
                <button onclick="navigator.clipboard.writeText('{licenca.chave_ativacao}')">Copiar Chave</button>
            </div>
            <hr>
            """
    else:
        html += "<p>Nenhuma licenÃ§a cadastrada!</p>"
    
    html += """
        <script>
            function copiarTodas() {
                let chaves = [];
                document.querySelectorAll('code').forEach(code => {
                    chaves.push(code.textContent);
                });
                navigator.clipboard.writeText(chaves.join('\\n'));
                alert('Todas as chaves copiadas!');
            }
        </script>
        <button onclick="copiarTodas()">Copiar Todas as Chaves</button>
    </body>
    </html>
    """
    
    return html



# ==================== RUN ====================

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        
        # CorreÃ§Ã£o automÃ¡tica: Adicionar coluna acesso_relatorios se nÃ£o existir
        try:
            from sqlalchemy import text
            with db.engine.connect() as conn:
                conn.execute(text('ALTER TABLE usuario ADD COLUMN acesso_relatorios BOOLEAN DEFAULT 0'))
                conn.commit()
            print("OK: Coluna 'acesso_relatorios' adicionada automaticamente!")
        except Exception:
            pass  # Coluna jÃ¡ existe ou erro ao adicionar

        # Criar admin se nÃ£o existir
        if not Usuario.query.filter_by(perfil='ADMIN').first():
            admin_password = (os.environ.get('ADMIN_PASSWORD') or '').strip() or secrets.token_urlsafe(12)
            admin = Usuario(
                nome='admin',
                senha=generate_password_hash(admin_password),
                perfil='ADMIN',
                ativo=True,
                acesso_configuracoes=True
            )
            db.session.add(admin)
            db.session.commit()
            print('✅ Usuário admin criado (ADMIN_PASSWORD não definido).')
            print('⚠️  Altere a senha imediatamente pelo painel de usuários.')
            print(f"🔐 Senha temporária (somente esta vez): {admin_password}")
            
        print("Banco atualizado!")

    def _porta_livre(porta):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            try:
                sock.bind(('0.0.0.0', porta))
                return True
            except OSError:
                return False

    def _selecionar_porta():
        candidatos = []
        for chave in ('APP_PORT', 'PORT'):
            valor = os.environ.get(chave)
            if valor:
                try:
                    candidatos.append(int(valor))
                except ValueError:
                    pass
        candidatos.extend([5000, 5001, 8000])
        vistos = set()
        for porta in candidatos:
            if porta in vistos:
                continue
            vistos.add(porta)
            if _porta_livre(porta):
                return porta
        return 5001

    porta = _selecionar_porta()
    if porta != 5000:
        print(f"Porta 5000 ocupada ou indisponivel. Usando {porta}.")

    print(f"Servidor Flask iniciado em http://127.0.0.1:{porta}")

    app.run(
        host='0.0.0.0',
        port=porta,
        debug=bool(os.environ.get('FLASK_DEBUG', '')),
        use_reloader=False
    )
